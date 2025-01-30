import streamlit as st
import json
import datetime
import re
import time
import io
import zipfile
import requests as req
import unicodedata
import pandas as pd

import gspread
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError

# gspread-formatting
from gspread_formatting import (
    format_cell_range,
    CellFormat,
    Color,
    TextFormat,
    set_column_width,
    set_row_height
)

from collections import defaultdict

# (신규) openpyxl
import openpyxl
from openpyxl import Workbook


# ========== [1] 인증/초기설정 =============
SCOPES = [
    "https://www.googleapis.com/auth/drive",
    "https://www.googleapis.com/auth/spreadsheets",
]

def get_credentials_from_secrets(which: str = "A") -> Credentials:
    """
    which="A"  -> st.secrets["google_service_account_a"] 사용
    which="B"  -> st.secrets["google_service_account_b"] 사용
    """
    if which.upper() == "A":
        service_account_info = st.secrets["google_service_account_a"]
    else:
        service_account_info = st.secrets["google_service_account_b"]

    credentials = Credentials.from_service_account_info(
        service_account_info,
        scopes=SCOPES
    )
    return credentials


# ----------------------------------------------------------------
# 검증(비교) 및 기타 헬퍼
# ----------------------------------------------------------------

def debug_hex(s: str) -> str:
    """문자열 s의 각 문자를 유니코드 코드포인트(\\uXXXX) 형태로 변환."""
    return " ".join(f"\\u{ord(ch):04X}" for ch in s)


def clean_artist_name(raw_name: str) -> str:
    """
    1) 유니코드 정규화(NFKC)
    2) 모든 제어문자(Category=C) 제거
    3) \xa0, \u3000 같은 특수 공백 치환
    4) strip()
    """
    import unicodedata
    import re

    if not raw_name:
        return ""

    # 1) 유니코드 정규화
    normalized = unicodedata.normalize('NFKC', raw_name)

    # 2) "모든 제어문자" 제거 (제어문자: Cc, Cf, Cs, Co, Cn 등)
    no_ctrl = "".join(ch for ch in normalized if not unicodedata.category(ch).startswith("C"))

    # 3) 특수공백 치환 + strip
    replaced = []
    for ch in no_ctrl:
        # ch가 Category=Z 공백인지
        if unicodedata.category(ch).startswith("Z"):
            replaced.append(" ")  # 전부 ' '로 치환
        else:
            replaced.append(ch)
    no_ctrl = "".join(replaced)

    # 4) 특수공백 치환
    no_ctrl = no_ctrl.replace('\xa0',' ').replace('\u3000',' ')

    # 5) strip
    cleaned = no_ctrl.strip()

    return cleaned

def show_detailed_verification():
    check_dict = st.session_state.get("check_dict", {})
    dv = check_dict.get("details_verification", {})
    if not dv:
        st.warning("세부 검증 데이터가 없습니다.")
        return

    tabA, tabB = st.tabs(["정산서 검증", "세부매출 검증"])

    with tabA:
        st.write("#### 정산서 검증")
        rows = dv.get("정산서", [])
        if not rows:
            st.info("정산서 검증 데이터가 없습니다.")
        else:
            import pandas as pd
            df = pd.DataFrame(rows)

            bool_cols = [c for c in df.columns if c.startswith("match_")]

            def highlight_boolean(val):
                if val is True:
                    return "background-color: #AAFFAA"
                elif val is False:
                    return "background-color: #FFAAAA"
                else:
                    return ""

            int_columns = [
                "원본_곡비", "정산서_곡비",
                "원본_공제금액", "정산서_공제금액",
                "원본_공제후잔액", "정산서_공제후잔액",
                "원본_정산율(%)", "정산서_정산율(%)"
            ]
            format_dict = {col: "{:.0f}" for col in int_columns if col in df.columns}

            st.dataframe(
                df.style
                  .format(format_dict)
                  .applymap(highlight_boolean, subset=bool_cols)
            )

    with tabB:
        st.write("#### 세부매출 검증")
        rows = dv.get("세부매출", [])
        if not rows:
            st.info("세부매출 검증 데이터가 없습니다.")
        else:
            import pandas as pd
            df = pd.DataFrame(rows)
            bool_cols = [c for c in df.columns if c.startswith("match_")]

            def highlight_boolean(val):
                if val is True:
                    return "background-color: #AAFFAA"
                elif val is False:
                    return "background-color: #FFAAAA"
                else:
                    return ""

            int_columns = ["원본_매출액", "정산서_매출액"]
            format_dict = {col: "{:.0f}" for col in int_columns if col in df.columns}

            st.dataframe(
                df.style
                  .format(format_dict)
                  .applymap(highlight_boolean, subset=bool_cols)
            )


def compare_artists(song_artists, revenue_artists):
    set_song = set(song_artists)
    set_revenue = set(revenue_artists)
    return {
        "missing_in_song": sorted(set_revenue - set_song),
        "missing_in_revenue": sorted(set_song - set_revenue),
        "common_count": len(set_song & set_revenue),
        "song_count": len(set_song),
        "revenue_count": len(set_revenue),
    }

def normalized_month(m):
    m = m.strip()
    if re.match(r'^\d{6}$', m):  # 202412
        yyyy = int(m[:4])
        mm = int(m[4:])
        return (yyyy, mm)
    pat = r'^(\d{4})년\s*(\d{1,2})월$'
    mmatch = re.match(pat, m)
    if mmatch:
        yyyy = int(mmatch.group(1))
        mm = int(mmatch.group(2))
        return (yyyy, mm)
    return m

def almost_equal(a, b, tol=1e-3):
    return abs(a - b) < tol

def get_next_month_str(ym: str) -> str:
    year = int(ym[:4])
    month = int(ym[4:])
    month += 1
    if month > 12:
        year += 1
        month = 1
    return f"{year}{month:02d}"

def get_prev_month_str(ym: str) -> str:
    """
    'YYYYMM' → 바로 직전 달 'YYYYMM'
    예) 202501 → 202412
    """
    year = int(ym[:4])
    month = int(ym[4:])
    month -= 1
    if month < 1:
        year -= 1
        month = 12
    return f"{year}{month:02d}"

def create_new_spreadsheet(filename: str, folder_id: str, drive_svc, attempt=1, max_attempts=5) -> str:
    try:
        query = (
            f"parents in '{folder_id}' and trashed=false "
            f"and name='{filename}'"
        )
        response = drive_svc.files().list(
            q=query,
            fields="files(id, name)",
            pageSize=50
        ).execute()
        files = response.get("files", [])
        if files:
            existing_file_id = files[0]["id"]
            print(f"파일 '{filename}' 이미 존재 -> 재사용 (ID={existing_file_id})")
            return existing_file_id

        file_metadata = {
            "name": filename,
            "mimeType": "application/vnd.google-apps.spreadsheet",
            "parents": [folder_id],
        }
        file = drive_svc.files().create(body=file_metadata).execute()
        return file["id"]

    except HttpError as e:
        if (e.resp.status == 403 and
            "userRateLimitExceeded" in str(e) and
            attempt < max_attempts):
            sleep_sec = 2 ** attempt
            print(f"[WARN] userRateLimitExceeded -> {sleep_sec}초 후 재시도 ({attempt}/{max_attempts})")
            time.sleep(sleep_sec)
            return create_new_spreadsheet(filename, folder_id, drive_svc, attempt+1, max_attempts)
        else:
            raise e

def batch_add_sheets(spreadsheet_id, sheet_svc, list_of_sheet_titles):
    meta = sheet_svc.spreadsheets().get(spreadsheetId=spreadsheet_id).execute()
    existing_sheets = meta["sheets"]
    existing_titles = [s["properties"]["title"] for s in existing_sheets]

    missing = [t for t in list_of_sheet_titles if t not in existing_titles]
    if not missing:
        print("모든 시트가 이미 존재합니다.")
        return

    BATCH_SIZE = 30
    requests_add = []
    total_count = 0
    
    for title in missing:
        requests_add.append({
            "addSheet": {
                "properties": {
                    "title": title,
                    "gridProperties": {
                        "rowCount": 200,
                        "columnCount": 8
                    }
                }
            }
        })

        if len(requests_add) >= BATCH_SIZE:
            body = {"requests": requests_add}
            resp = sheet_svc.spreadsheets().batchUpdate(
                spreadsheetId=spreadsheet_id,
                body=body
            ).execute()

            total_count += len(resp["replies"])
            print(f"분할 addSheet 완료: {len(resp['replies'])}개 생성")
            requests_add.clear()
            time.sleep(2)

    if requests_add:
        body = {"requests": requests_add}
        resp = sheet_svc.spreadsheets().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body=body
        ).execute()
        total_count += len(resp["replies"])
        print(f"마지막 addSheet 완료: {len(resp['replies'])}개 생성")
        requests_add.clear()

    print(f"시트 생성 총 개수: {total_count}")
    for idx, rep in enumerate(resp["replies"]):
        sheet_props = rep["addSheet"]["properties"]
        print(f" -> {idx} '{sheet_props['title']}' (sheetId={sheet_props['sheetId']})")

def duplicate_worksheet_with_new_name(gs_obj, from_sheet_name: str, to_sheet_name: str):
    all_ws = gs_obj.worksheets()
    all_titles = [w.title for w in all_ws]
    from_ws = None
    for w in all_ws:
        if w.title == from_sheet_name:
            from_ws = w
            break
    if not from_ws:
        raise ValueError(f"원본 시트 '{from_sheet_name}'를 찾을 수 없습니다.")

    base_name = to_sheet_name
    idx = 2
    while to_sheet_name in all_titles:
        to_sheet_name = f"{base_name} ({idx})"
        idx += 1

    new_ws = gs_obj.duplicate_sheet(source_sheet_id=from_ws.id, new_sheet_name=to_sheet_name)
    return new_ws

def is_korean_char(ch: str):
    return "가" <= ch <= "힣"

def is_korean_string(s: str):
    return any(is_korean_char(ch) for ch in s)

def album_sort_key(album_name: str):
    return (0 if is_korean_string(album_name) else 1, album_name)

def to_currency(num):
    return f"₩{format(int(round(num)), ',')}"

def update_next_month_tab(song_cost_sh, ym: str):
    """
    예시 함수 (기존 코드 내 사용)
    """
    old_ws = song_cost_sh.worksheet(ym)
    old_data = old_ws.get_all_values()
    if not old_data:
        print(f"'{ym}' 탭이 비어 있음")
        return

    old_header = old_data[0]
    old_body   = old_data[1:]

    try:
        idx_artist_old = old_header.index("아티스트명")
        idx_remain_old = old_header.index("당월 잔액")
    except ValueError:
        print("이전 달 시트에 '아티스트명' 또는 '당월 잔액' 칼럼이 없습니다.")
        return

    # 전월 잔액을 dict로 모아둠
    prev_month_dict = {}
    for row in old_body:
        artist_name = row[idx_artist_old].strip()
        if not artist_name or artist_name in ("합계","총계"):
            continue
        try:
            remain_val = float(row[idx_remain_old].replace(",", ""))
        except:
            remain_val = 0.0
        prev_month_dict[artist_name] = remain_val

    # 다음 달 시트 만들기(복제)
    next_ym = get_next_month_str(ym)
    new_ws = duplicate_worksheet_with_new_name(song_cost_sh, ym, next_ym)
    
    # 복제된 시트의 데이터 읽기
    new_data = new_ws.get_all_values()
    if not new_data:
        print(f"복제된 '{next_ym}' 탭이 비어 있습니다.")
        return

    new_header = new_data[0]
    try:
        idx_artist_new = new_header.index("아티스트명")
        idx_prev_new   = new_header.index("전월 잔액")
        idx_deduct_new = new_header.index("당월 차감액")
        # idx_remain_new = new_header.index("당월 잔액")
    except ValueError:
        print("새로 만든 시트(다음 달 탭)에 필요한 칼럼이 없습니다.")
        return
    
    # 본문 (마지막 합계 행은 제외)
    content = new_data[1:-1]

    updated_prev_vals = []   # D열에 들어갈 값
    updated_deduct_vals = [] # F열에 들어갈 값

    for row in content:
        artist = row[idx_artist_new].strip()
        old_val = prev_month_dict.get(artist, 0.0)  # 전월 잔액
        updated_prev_vals.append([old_val])
        updated_deduct_vals.append(["0"])  # 당월 차감액은 0으로 초기화

    row_count = len(content)
    start_row = 2
    end_row   = 1 + row_count

    # batch_update에 쓸 requests
    requests_body = [
        {
            "range": f"F{start_row}:F{end_row}",
            "values": updated_prev_vals
        },
        {
            "range": f"G{start_row}:G{end_row}",
            "values": updated_deduct_vals
        }
    ]

    # 한 번에 batch_update로 호출
    new_ws.batch_update(
        requests_body,
        value_input_option="USER_ENTERED"
    )

    print(f"'{ym}' → '{next_ym}' 탭 복제 및 전월/당월 차감액만 갱신(배치 업데이트) 완료!")


def is_summary_row(cleaned_artist_name: str) -> bool:
    """
    아티스트명이 공란('')이거나,
    '합계', '총계', 'TOTAL', 'total' 같은 문자열이면
    합계행으로 간주해서 True 리턴
    """
    if not cleaned_artist_name:
        return True
    # 대소문자 구분 없이 모두 upper() 해서 비교
    up = cleaned_artist_name.upper()
    return (up in ("합계", "총계", "TOTAL"))



# ------------------------------------------------------------------------------
# (A) "0) 곡비 파일 수정" 섹션
# ------------------------------------------------------------------------------
def section_zero_prepare_song_cost():
    """
    - 이번 달(YYYYMM)과 직전 달(YYYYMM) 탭을 열어, 
      '전월 잔액 + 당월 발생액' vs (UMAG + FLUXUS매출) 비교 → '당월 차감액' 갱신
    - 소속이 여러 개인 경우에도 (UMAG + FLUXUS) 매출 모두 합산
    """
    import pandas as pd  # 함수 시작부

    st.subheader("0) 곡비 파일 수정")

    default_ym = st.session_state.get("ym", "")
    new_ym = st.text_input("진행기간(YYYYMM) - (곡비 파일 수정용)", default_ym)

    if st.button("곡비 파일 수정하기"):
        creds_a = get_credentials_from_secrets("A")
        gc_a = gspread.authorize(creds_a)

        if not re.match(r'^\d{6}$', new_ym):
            st.error("진행기간은 YYYYMM 6자리로 입력해야 합니다.")
            return

        st.session_state["ym"] = new_ym
        prev_ym = get_prev_month_str(new_ym)

        # (1) input_song cost 열기
        try:
            song_cost_sh = gc_a.open("input_song cost")
        except gspread.exceptions.SpreadsheetNotFound:
            st.error("Google Sheet 'input_song cost'를 찾을 수 없습니다.")
            return

        # (2) umag / fluxus_song / fluxus_yt 열기
        try:
            umag_sh = gc_a.open("input_online revenue_umag_integrated")
        except:
            st.error("'input_online revenue_umag_integrated' 없음")
            return
        
        try:
            fluxus_song_sh = gc_a.open("input_online revenue_fluxus_song")
        except:
            st.error("'input_online revenue_fluxus_song' 없음")
            return
        
        try:
            fluxus_yt_sh = gc_a.open("input_online revenue_fluxus_yt")
        except:
            st.error("'input_online revenue_fluxus_yt' 없음")
            return

        # ---------------------------
        # 0-A) 직전 달(YYYYMM) 탭에서 '아티스트별 당월 잔액' dict
        # ---------------------------
        ws_map_sc = {ws.title: ws for ws in song_cost_sh.worksheets()}
        if prev_ym not in ws_map_sc:
            st.error(f"'input_song cost'에 직전 달 '{prev_ym}' 탭이 없습니다.")
            return

        ws_prev = ws_map_sc[prev_ym]
        data_prev = ws_prev.get_all_values()
        if not data_prev:
            st.error(f"'{prev_ym}' 탭이 비어있음")
            return

        header_prev = data_prev[0]
        body_prev = data_prev[1:]
        try:
            idx_artist_p = header_prev.index("아티스트명")
            idx_remain_p = header_prev.index("당월 잔액")
        except ValueError as e:
            st.error(f"직전 달 '{prev_ym}' 시트에 '아티스트명' 또는 '당월 잔액' 없음: {e}")
            return

        prev_remain_dict = {}
        for row_p in body_prev:
            artist_p = clean_artist_name(row_p[idx_artist_p])
            if not artist_p or artist_p in ("합계","총계"):
                continue
            try:
                val_p = float(row_p[idx_remain_p].replace(",",""))
            except:
                val_p = 0.0
            prev_remain_dict[artist_p] = val_p

        # ---------------------------
        # 0-B) 이번 달(YYYYMM) 탭 read
        # ---------------------------
        if new_ym not in ws_map_sc:
            st.error(f"이번 달 '{new_ym}' 탭이 없습니다.")
            return

        ws_new = ws_map_sc[new_ym]
        data_new = ws_new.get_all_values()
        if not data_new:
            st.error(f"'{new_ym}' 탭이 비어있음")
            return

        header_new = data_new[0]
        body_new   = data_new[1:-1]  # 마지막 합계행 제외

        try:
            idx_sosok_n  = header_new.index("소속")
            idx_artist_n = header_new.index("아티스트명")
            idx_prev_n   = header_new.index("전월 잔액")
            idx_curr_n   = header_new.index("당월 발생액")
            idx_ded_n    = header_new.index("당월 차감액")
        except:
            st.error(f"[{new_ym}] 탭에 '소속' 또는 '당월 차감액' 등이 없음")
            return

        # ---------------------------
        # 0-C) UMAG 인풋 read
        # ---------------------------
        ws_map_umag = {ws.title: ws for ws in umag_sh.worksheets()}
        if new_ym not in ws_map_umag:
            st.error(f"'input_online revenue_umag_integrated'에 '{new_ym}' 탭 없음")
            return
        ws_umag = ws_map_umag[new_ym]
        data_umag = ws_umag.get_all_values()
        header_umag = data_umag[0]
        body_umag   = data_umag[1:]

        # [추가] 합계행(요약행) 필터링
        col_artist_umag = header_umag.index("앨범아티스트")
        filtered_umag = []
        for row_u in body_umag:
            a_u = clean_artist_name(row_u[col_artist_umag])
            if is_summary_row(a_u):  
                # 합계/총계/공란 → skip
                continue
            filtered_umag.append(row_u)

        body_umag = filtered_umag

        try:
            col_artist_umag  = header_umag.index("앨범아티스트")
            col_revenue_umag = header_umag.index("권리사정산금액")
        except:
            st.error("'앨범아티스트' / '권리사정산금액' 칼럼 필요(UMAG)")
            return

        from collections import defaultdict
        sum_umag_dict = defaultdict(float)
        for row_u in body_umag:
            a_u = clean_artist_name(row_u[col_artist_umag])

            # 소문자로 변환한 값
            a_lower = a_u.lower()
            # (1) 아티스트명이 공란, (2) 아티스트명 안에 '합계','총계','total' 포함, 
            # (3) 전부 숫자인 경우 -> 합계행으로 보고 스킵
            if (
                not a_u 
                or '합계' in a_u 
                or '총계' in a_u 
                or 'total' in a_lower 
                or a_u.isdigit()
            ):
                continue

            try:
                val_u = float(row_u[col_revenue_umag].replace(",",""))
            except:
                val_u = 0.0
            sum_umag_dict[a_u] += val_u

        # ---------------------------
        # 0-D) fluxus_song read
        # ---------------------------
        ws_map_flux_song = {ws.title: ws for ws in fluxus_song_sh.worksheets()}
        if new_ym not in ws_map_flux_song:
            st.error(f"'fluxus_song'에 '{new_ym}' 탭이 없음")
            return
        ws_fs = ws_map_flux_song[new_ym]
        data_fs = ws_fs.get_all_values()
        if not data_fs:
            st.error(f"'{new_ym}' 탭(fluxus_song) 비어있음")
            return
        header_fs = data_fs[0]
        body_fs   = data_fs[1:]

        # [추가] 합계/요약행 필터링
        col_artist_fs = header_fs.index("가수명")
        filtered_fs = []
        for row_fs_ in body_fs:
            artist_temp = clean_artist_name(row_fs_[col_artist_fs])
            if is_summary_row(artist_temp):
                continue
            filtered_fs.append(row_fs_)
        body_fs = filtered_fs

        try:
            col_artist_fs = header_fs.index("가수명")
            col_revenue_fs= header_fs.index("권리사 정산액")
        except:
            st.error("fluxus_song: '가수명' / '권리사 정산액' 칼럼 필요")
            return

        sum_flux_song_dict = defaultdict(float)
        for row_fs in body_fs:
            a_fs = clean_artist_name(row_fs[col_artist_fs])

            # 소문자로 변환한 값
            a_lower = a_u.lower()
            # (1) 아티스트명이 공란, (2) 아티스트명 안에 '합계','총계','total' 포함, 
            # (3) 전부 숫자인 경우 -> 합계행으로 보고 스킵
            if (
                not a_u 
                or '합계' in a_u 
                or '총계' in a_u 
                or 'total' in a_lower 
                or a_u.isdigit()
            ):
                continue

            try:
                val_fs = float(row_fs[col_revenue_fs].replace(",",""))
            except:
                val_fs = 0.0
            sum_flux_song_dict[a_fs] += val_fs

        # ---------------------------
        # 0-E) fluxus_yt read
        # ---------------------------
        ws_map_flux_yt = {ws.title: ws for ws in fluxus_yt_sh.worksheets()}
        if new_ym not in ws_map_flux_yt:
            st.error(f"'fluxus_yt'에 '{new_ym}' 탭 없음")
            return
        ws_fy = ws_map_flux_yt[new_ym]
        data_fy = ws_fy.get_all_values()
        if not data_fy:
            st.error(f"'{new_ym}' 탭(fluxus_yt) 비어있음")
            return
        header_fy = data_fy[0]
        body_fy   = data_fy[1:]

        # [추가] 합계/요약행 필터링
        col_artist_fy = header_fy.index("ALBIM ARTIST")
        filtered_fy = []
        for row_fy_ in body_fy:
            artist_temp = clean_artist_name(row_fy_[col_artist_fy])
            if is_summary_row(artist_temp):
                continue
            filtered_fy.append(row_fy_)
        body_fy = filtered_fy

        try:
            col_artist_fy  = header_fy.index("ALBIM ARTIST")
            col_revenue_fy = header_fy.index("권리사 정산액 \n(KRW)")
        except:
            st.error("'fluxus_yt' 칼럼( ALBIM ARTIST, 권리사 정산액 \n(KRW) ) 필요")
            return

        sum_flux_yt_dict = defaultdict(float)
        for row_fy in body_fy:
            a_fy = clean_artist_name(row_fy[col_artist_fy])

            # 소문자로 변환한 값
            a_lower = a_u.lower()
            # (1) 아티스트명이 공란, (2) 아티스트명 안에 '합계','총계','total' 포함, 
            # (3) 전부 숫자인 경우 -> 합계행으로 보고 스킵
            if (
                not a_u 
                or '합계' in a_u 
                or '총계' in a_u 
                or 'total' in a_lower 
                or a_u.isdigit()
            ):
                continue

            try:
                val_fy = float(row_fy[col_revenue_fy].replace(",",""))
            except:
                val_fy = 0.0
            sum_flux_yt_dict[a_fy] += val_fy

        # ---------------------------------------
        # [중요] 2개 이상 소속도 “모두” 매출 더해서 actual_deduct 산출
        # ---------------------------------------
        updated_vals_for_def = []

        for row_idx, row_data in enumerate(body_new):
            artist_n = clean_artist_name(row_data[idx_artist_n])
            if not artist_n or artist_n in ("합계","총계"):
                updated_vals_for_def.append(["","",""])
                continue

            sosok_str = row_data[idx_sosok_n].strip().upper()
            affils = re.split(r'[,&/]', sosok_str)  # "UMAG,FLUXUS" → ["UMAG","FLUXUS"]
            affils = [x.strip() for x in affils if x.strip()]

            try:
                curr_val_str = row_data[idx_curr_n].replace(",","")
                curr_val = float(curr_val_str) if curr_val_str else 0.0
            except:
                curr_val = 0.0

            prev_val = prev_remain_dict.get(artist_n,0.0)

            # 모든 소속매출 합산
            total_revenue = 0.0
            for one_sosok in affils:
                if one_sosok == "UMAG":
                    total_revenue += sum_umag_dict.get(artist_n, 0.0)
                elif one_sosok == "FLUXUS":
                    fs_val = sum_flux_song_dict.get(artist_n, 0.0)
                    fy_val = sum_flux_yt_dict.get(artist_n, 0.0)
                    total_revenue += (fs_val + fy_val)
                else:
                    # 알수없는 소속? pass
                    pass

            can_deduct = prev_val + curr_val
            actual_deduct = min(total_revenue, can_deduct)

            updated_vals_for_def.append([prev_val, curr_val, actual_deduct])

        # batch_update → (E:F:G) or (D:E:F) 등 실제 칼럼 위치 맞춤
        total_rows = len(body_new)
        start_row = 2
        end_row   = 1 + total_rows
        range_notation = f"E{start_row}:G{end_row}"  # (전월/당월발생/당월차감)
        requests_body = [{"range": range_notation, "values": updated_vals_for_def}]
        ws_new.batch_update(requests_body, value_input_option="USER_ENTERED")

        #--------------------------------
        # 아티스트 수 검증
        #--------------------------------
        umag_count_artists = 0
        fluxus_count_artists = 0

        # 곡비파일(body_new)에서 소속을 보고 카운팅
        for row_data in body_new:
            sosok_str = row_data[idx_sosok_n].strip().upper()
            affils = re.split(r'[,&/]', sosok_str)
            affils = [x.strip() for x in affils if x.strip()]

            # 소속 문자열 안에 "UMAG"가 있으면 UMAG 카운트
            if "UMAG" in affils:
                umag_count_artists += 1

            # 소속 문자열 안에 "FLUXUS"가 있으면 FLUXUS 카운트
            if "FLUXUS" in affils:
                fluxus_count_artists += 1

        # 매출 인풋파일들의 "원본" 행 개수
        umag_raw_rows = len(body_umag)   # 예: UMAG 매출
        flux_song_raw_rows = len(body_fs) # fluxus_song
        flux_yt_raw_rows   = len(body_fy) # fluxus_yt

        # 사용자 안내용
        st.session_state["verification_original"] = {
            "곡비파일": {
                "UMAG_아티스트수": umag_count_artists,
                "FLUXUS_아티스트수": fluxus_count_artists
            },
            "매출액파일": {
                "UMAG행개수": umag_raw_rows,
                "FLUXUS_SONG행개수": flux_song_raw_rows,
                "FLUXUS_YT행개수": flux_yt_raw_rows
            }
        }

        # --------------------------------
        # A) 소속별 "아티스트 set" 구성
        # --------------------------------
        umag_artists_from_cost = set()
        fluxus_artists_from_cost = set()

        for row_data in body_new:
            artist_n = clean_artist_name(row_data[idx_artist_n])
            sosok_str = row_data[idx_sosok_n].strip().upper()
            affils = re.split(r'[,&/]', sosok_str)
            affils = [x.strip() for x in affils if x.strip()]

            # 소속 중에 UMAG가 하나라도 있으면 => umag_artists_from_cost.add(artist_n)
            if "UMAG" in affils:
                umag_artists_from_cost.add(artist_n)

            # 소속 중에 FLUXUS가 하나라도 있으면 => fluxus_artists_from_cost.add(artist_n)
            if "FLUXUS" in affils:
                fluxus_artists_from_cost.add(artist_n)


        # 2) UMAG 인풋파일 '누락행' 탐색
        missing_umag_rows = []
        umag_processed_rows = 0

        for i, row_u in enumerate(body_umag):
            raw_artist = row_u[col_artist_umag]
            cleaned_a = clean_artist_name(raw_artist)
            if cleaned_a in umag_artists_from_cost:
                # 처리됨
                umag_processed_rows += 1
            else:
                # 누락
                missing_umag_rows.append({
                    "row_idx": i,
                    "raw_artist": raw_artist,
                    "cleaned_artist": cleaned_a,
                    "reason": "곡비파일에 아티스트 없음(UMAG_INTEGRATED)"
                })

        # 3) fluxus_song 누락행
        missing_flux_song_rows = []
        fluxus_song_processed_rows = 0
        for i, row_fs in enumerate(body_fs):
            raw_artist = row_fs[col_artist_fs]
            cleaned_a = clean_artist_name(raw_artist)
            if cleaned_a in fluxus_artists_from_cost:
                fluxus_song_processed_rows += 1
            else:
                missing_flux_song_rows.append({
                    "row_idx": i,
                    "raw_artist": raw_artist,
                    "cleaned_artist": cleaned_a,
                    "reason": "곡비파일에 아티스트 없음(FLUXUS_SONG)"
                })

        # 4) fluxus_yt 누락행
        missing_flux_yt_rows = []
        fluxus_yt_processed_rows = 0
        for i, row_fy in enumerate(body_fy):
            raw_artist = row_fy[col_artist_fy]
            cleaned_a = clean_artist_name(raw_artist)
            if cleaned_a in fluxus_artists_from_cost:
                fluxus_yt_processed_rows += 1
            else:
                missing_flux_yt_rows.append({
                    "row_idx": i,
                    "raw_artist": raw_artist,
                    "cleaned_artist": cleaned_a,
                    "reason": "곡비파일에 아티스트 없음(FLUXUS_YT)"
                })

        st.session_state["missing_rows"] = {
            "UMAG": missing_umag_rows,
            "FLUXUS_SONG": missing_flux_song_rows,
            "FLUXUS_YT": missing_flux_yt_rows
        }


        # --------------------------------
        # B) 처리된 아티스트 수
        # --------------------------------
        umag_count_processed = len(umag_artists_from_cost)
        fluxus_count_processed = len(fluxus_artists_from_cost)

        # --------------------------------
        # C) 실제 매출 행 처리 개수
        # --------------------------------

        # (1) UMAG
        umag_processed_rows = 0
        for row_u in body_umag:
            raw_artist = row_u[col_artist_umag]
            a = clean_artist_name(raw_artist)
            if a in umag_artists_from_cost:  # 곡비에도 있고, 소속=UMAG인 아티스트
                umag_processed_rows += 1

        # (2) FLUXUS SONG
        fluxus_song_processed_rows = 0
        for row_fs in body_fs:
            raw_artist = row_fs[col_artist_fs]
            a = clean_artist_name(raw_artist)
            if a in fluxus_artists_from_cost:
                fluxus_song_processed_rows += 1

        # (3) FLUXUS YT
        fluxus_yt_processed_rows = 0
        for row_fy in body_fy:
            raw_artist = row_fy[col_artist_fy]
            a = clean_artist_name(raw_artist)
            if a in fluxus_artists_from_cost:
                fluxus_yt_processed_rows += 1

        # --------------------------------
        # D) st.session_state 저장
        # --------------------------------
        st.session_state["verification_processed"] = {
            "곡비파일": {
                "UMAG_아티스트수": umag_count_processed,
                "FLUXUS_아티스트수": fluxus_count_processed
            },
            "매출액파일": {
                "UMAG행개수": umag_processed_rows,
                "FLUXUS_SONG행개수": fluxus_song_processed_rows,
                "FLUXUS_YT행개수": fluxus_yt_processed_rows
            }
        }


        #--------------------------------
        # 검증 결과 출력
        #--------------------------------
        # 곡비파일 수정 완료 후 검증 결과를 보여주는 영역
        st.write("### 검증 결과")

        tab_summary, tab_missing = st.tabs(["검증 요약","누락 행 목록"])

        if "verification_original" in st.session_state and "verification_processed" in st.session_state:
            orig = st.session_state["verification_original"]
            proc = st.session_state["verification_processed"]

        with tab_summary:    
            # 여기서 곡비파일 아티스트 수 / 매출액 행 개수 비교
            st.write("#### 곡비파일 아티스트 검증")
            st.write(f"- (원본) UMAG: {orig['곡비파일']['UMAG_아티스트수']}, FLUXUS: {orig['곡비파일']['FLUXUS_아티스트수']}")
            st.write(f"- (처리) UMAG: {proc['곡비파일']['UMAG_아티스트수']}, FLUXUS: {proc['곡비파일']['FLUXUS_아티스트수']}")

            diff_umag_artist = (orig["곡비파일"]["UMAG_아티스트수"] - proc["곡비파일"]["UMAG_아티스트수"])
            diff_flux_artist = (orig["곡비파일"]["FLUXUS_아티스트수"] - proc["곡비파일"]["FLUXUS_아티스트수"])

            # 차이가 있으면 메시지
            if diff_umag_artist != 0 or diff_flux_artist != 0:
                st.warning(f"아티스트 수에 차이가 발생했습니다! (UMAG: {diff_umag_artist}, FLUXUS: {diff_flux_artist})")

            st.write("#### 매출액 행 개수 검증")
            st.write(f"- (원본) UMAG: {orig['매출액파일']['UMAG행개수']}, FLUXUS_SONG: {orig['매출액파일']['FLUXUS_SONG행개수']}, FLUXUS_YT: {orig['매출액파일']['FLUXUS_YT행개수']}")
            st.write(f"- (처리) UMAG: {proc['매출액파일']['UMAG행개수']}, FLUXUS_SONG: {proc['매출액파일']['FLUXUS_SONG행개수']}, FLUXUS_YT: {proc['매출액파일']['FLUXUS_YT행개수']}")

            diff_umag_row = orig["매출액파일"]["UMAG행개수"] - proc["매출액파일"]["UMAG행개수"]
            diff_flux_song = orig["매출액파일"]["FLUXUS_SONG행개수"] - proc["매출액파일"]["FLUXUS_SONG행개수"]
            diff_flux_yt   = orig["매출액파일"]["FLUXUS_YT행개수"] - proc["매출액파일"]["FLUXUS_YT행개수"]

            if diff_umag_row!=0 or diff_flux_song!=0 or diff_flux_yt!=0:
                st.warning(f"매출 데이터 행개수 차이 발생!")
                st.warning(f"UMAG: {diff_umag_row}개    /   FLUXUS_SONG: {diff_flux_song}개    /   FLUXUS_YT: {diff_flux_yt}개")

            if diff_umag_artist==0 and diff_flux_artist==0 and diff_umag_row==0 and diff_flux_song==0 and diff_flux_yt==0:
                st.success("원본과 처리 결과가 모두 일치합니다!")
            else:
                st.error("원본 vs 처리 결과에 차이가 있습니다. 상세탭에서 상세 누락 행을 확인해 주세요.")

        with tab_missing:
            # 여기에서 UMAG / Fluxus Song / Fluxus YT ‘missing_rows’ 표
            if diff_umag_row!=0 or diff_flux_song!=0 or diff_flux_yt!=0:
                st.warning(f"매출 데이터 행개수 차이 발생!")
                st.warning(f"UMAG: {diff_umag_row}개    /   FLUXUS_SONG: {diff_flux_song}개    /   FLUXUS_YT: {diff_flux_yt}개")
                if "missing_rows" in st.session_state:
                    missing_all = st.session_state["missing_rows"]

                    # UMAG
                    if missing_all["UMAG"]:
                        st.write("#### 매출액_UMAG 누락 행 목록")
                        import pandas as pd
                        df_umag_miss = pd.DataFrame(missing_all["UMAG"])
                        st.dataframe(df_umag_miss)

                    # FLUXUS_SONG
                    if missing_all["FLUXUS_SONG"]:
                        st.write("#### 매출액_Fluxus_Song 누락 행 목록")
                        df_fs_miss = pd.DataFrame(missing_all["FLUXUS_SONG"])
                        st.dataframe(df_fs_miss)
                    
                    # FLUXUS_YT
                    if missing_all["FLUXUS_YT"]:
                        st.write("#### 매출액_Fluxus_YT 누락 행 목록")
                        df_fy_miss = pd.DataFrame(missing_all["FLUXUS_YT"])
                        st.dataframe(df_fy_miss)

        st.success(f"곡비 파일('{new_ym}' 탭) 수정 완료!")
        st.session_state["song_cost_prepared"] = True


# ------------------------------------------------------------------------------
# (B) "1) 정산 보고서 정보 입력 항목" 섹션
# ------------------------------------------------------------------------------
def section_one_report_input():
    """
    UMAG 아티스트: 기존과 동일한 형식
    FLUXUS 아티스트: 정산서와 세부매출내역의 칼럼 구조를 변경
    - (정산서) -> [앨범, 내용, 기간, 매출액]
    - (세부매출내역) -> [앨범아티스트, 앨범명, 트랙 No., 트랙명, 매출 순수익]
    - 중복소속(UMAG+FLUXUS)인 아티스트는 제외.
    """
    st.subheader("1) 정산 보고서 정보 입력 항목")

    default_ym = st.session_state.get("ym", "")
    default_report_date = st.session_state.get("report_date", "")

    ym = st.text_input("진행기간(YYYYMM)", default_ym)
    report_date = st.text_input("보고서 발행 날짜 (YYYY-MM-DD)", default_report_date)

    # 진행 상황 표시용
    progress_bar = st.empty()
    artist_placeholder = st.empty()

    if st.button("정산 보고서 생성 시작"):
        if not re.match(r'^\d{6}$', ym):
            st.error("진행기간은 YYYYMM 6자리로 입력하세요.")
            return
        if not report_date:
            st.error("보고서 발행 날짜를 입력하세요.")
            return

        st.session_state["ym"] = ym
        st.session_state["report_date"] = report_date

        creds_a = get_credentials_from_secrets("A")
        gc_a = gspread.authorize(creds_a)
        drive_svc_a = build("drive", "v3", credentials=creds_a)
        sheet_svc_a = build("sheets", "v4", credentials=creds_a)

        check_dict = {}
        out_file_id = generate_report(
            ym=ym,
            report_date=report_date,
            check_dict=check_dict,
            gc=gc_a,
            drive_svc=drive_svc_a,
            sheet_svc=sheet_svc_a,
            progress_bar=progress_bar,
            artist_placeholder=artist_placeholder
        )

        st.session_state["report_done"] = True
        st.session_state["report_file_id"] = out_file_id
        st.session_state["check_dict"] = check_dict

        st.success(f"보고서 생성 완료! file_id={out_file_id}")


# ------------------------------------------------------------------------------
# (C) 보고서 링크 & 검증
# ------------------------------------------------------------------------------
def section_two_sheet_link_and_verification():
    if "report_done" in st.session_state and st.session_state["report_done"]:
        st.subheader("2) 정산 보고서 시트링크 및 검증")

        tab1, tab2 = st.tabs(["보고서 링크 / 요약", "세부 검증 내용"])

        with tab1:
            out_file_id = st.session_state.get("report_file_id", "")
            if out_file_id:
                gsheet_url = f"https://docs.google.com/spreadsheets/d/{out_file_id}/edit"
                st.write(f"**생성된 구글시트 링크:** {gsheet_url}")

            # 중복 소속 아티스트 안내
            double_sosok_list = st.session_state.get("excluded_double_sosok", [])
            if double_sosok_list:
                st.warning(f"2개 소속 중복으로 작업 제외된 아티스트 목록: {double_sosok_list}")
                st.info("해당 아티스트는 '2개 소속이 되어 진행이 되지 않았습니다' 문구가 노출됩니다.")

        with tab2:
            st.write("### 세부 검증 내용")
            show_detailed_verification()

    else:
        st.info("정산 보고서 생성 완료 후 확인 가능합니다.")


# ------------------------------------------------------------------------------
# (D) 엑셀 업로드 → 아티스트별 XLSX 파일 분할
# ------------------------------------------------------------------------------
def section_three_upload_and_split_excel():
    """
    곡비-보고서 통합 완료 후, 최종 생성된 구글시트(Excel 다운로드본)를
    다시 업로드하여 아티스트별 XLSX로 분할 & ZIP으로 다운로드
    """
    if "report_done" not in st.session_state or not st.session_state["report_done"]:
        st.info("정산 보고서가 먼저 생성된 뒤에, 엑셀 업로드 가능합니다.")
        return

    st.subheader("3) 엑셀 업로드 후 [아티스트별] 엑셀파일로 분할 (서식 유지)")

    st.write("""
    **사용 순서**  
    1. 생성된 구글시트(보고서)에서 "**파일 → 다운로드 → Microsoft Excel (.xlsx)**"로 다운로드  
    2. 아래 업로드 버튼으로 방금 받은 .xlsx 파일을 업로드  
    3. 아티스트별로 '정산서' 탭 + '세부매출내역' 탭만 묶은 XLSX를 각각 만들고, ZIP으로 묶어 다운로드
    """)

    uploaded_file = st.file_uploader("엑셀 파일 업로드", type=["xlsx"])
    if uploaded_file is None:
        return

    progress_bar = st.progress(0.0)
    progress_text = st.empty()

    original_file_data = uploaded_file.read()
    try:
        wb_all = openpyxl.load_workbook(io.BytesIO(original_file_data))
    except Exception as e:
        st.error(f"엑셀 파일을 읽는 중 오류 발생: {e}")
        return

    sheet_names = wb_all.sheetnames

    # 아티스트별 (정산서, 세부매출내역) 탭 매칭
    from collections import defaultdict
    all_pairs = defaultdict(lambda: {"report": None, "detail": None})

    for sn in sheet_names:
        if sn.endswith("(정산서)"):
            artist_name = sn[:-5].strip()  # '(정산서)' 제거
            if artist_name.startswith("FLUXUS_"):
                artist_name = artist_name[len("FLUXUS_"):]
            elif artist_name.startswith("UMAG_"):
                artist_name = artist_name[len("UMAG_"):]
            else:
                st.error(f"알 수 없는 소속 형식: {artist_name}")
                continue
            all_pairs[artist_name]["report"] = sn
        elif sn.endswith("(세부매출내역)"):
            artist_name = sn[:-7].strip()  # '(세부매출내역)' 제거
            if artist_name.startswith("FLUXUS_"):
                artist_name = artist_name[len("FLUXUS_"):]
            elif artist_name.startswith("UMAG_"):
                artist_name = artist_name[len("UMAG_"):]
            else:
                st.error(f"알 수 없는 소속 형식: {artist_name}")
                continue
            all_pairs[artist_name]["detail"] = sn
        else:
            pass  # 무시


    # 소속 정보를 위해 input_song cost에서 다시 가져올 수도 있음
    # 여기서는 단순 예시
    artist_sosok_dict = {}
    try:
        creds_a = get_credentials_from_secrets("A")
        gc_a = gspread.authorize(creds_a)
        sc_sh = gc_a.open("input_song cost")
        sc_ws_map = {ws.title: ws for ws in sc_sh.worksheets()}
        current_ym = st.session_state.get("ym", "")
        if current_ym in sc_ws_map:
            ws_sc = sc_ws_map[current_ym]
            data_sc = ws_sc.get_all_values()
            if data_sc:
                hdr_sc = data_sc[0]
                try:
                    idx_a = hdr_sc.index("아티스트명")
                    idx_s = hdr_sc.index("소속")
                except:
                    idx_a, idx_s = -1, -1

                for row_sc in data_sc[1:]:
                    a = row_sc[idx_a].strip() if idx_a>=0 else ""
                    s = row_sc[idx_s].strip().upper() if idx_s>=0 else ""
                    if a and s:
                        artist_sosok_dict[a] = s
    except:
        pass

    # ZIP
    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        total_pairs = len(all_pairs)
        for i, (artist, pair_info) in enumerate(all_pairs.items()):
            ratio = (i+1)/total_pairs
            progress_bar.progress(ratio)
            progress_text.info(f"{int(ratio*100)}% - '{artist}' 처리 중...")

            ws_report_name = pair_info["report"]
            ws_detail_name = pair_info["detail"]
            if not ws_report_name or not ws_detail_name:
                # 한쪽 탭만 있는 경우는 skip
                continue

            sosok = artist_sosok_dict.get(artist, "UNKNOWN")

            # 새 워크북
            temp_wb = openpyxl.Workbook()
            def_ws = temp_wb.active
            temp_wb.remove(def_ws)

            # (1) 정산서 복사
            orig_ws_report = wb_all[ws_report_name]
            new_ws_report = temp_wb.create_sheet(ws_report_name)
            copy_sheet(orig_ws_report, new_ws_report)

            # (2) 세부매출내역 복사
            orig_ws_detail = wb_all[ws_detail_name]
            new_ws_detail = temp_wb.create_sheet(ws_detail_name)
            copy_sheet(orig_ws_detail, new_ws_detail)

            # 파일명 "소속_정산보고서_아티스트명_YYYYMM.xlsx"
            current_ym = st.session_state.get("ym", "000000")
            safe_artist = artist.replace("/", "_").replace("\\", "_")
            filename_xlsx = f"{sosok}_정산보고서_{safe_artist}_{current_ym}.xlsx"

            single_buf = io.BytesIO()
            temp_wb.save(single_buf)
            single_buf.seek(0)
            zf.writestr(filename_xlsx, single_buf.getvalue())

    zip_buf.seek(0)
    progress_text.success("아티스트별 엑셀 생성 완료!")
    st.download_button(
        label="ZIP 다운로드",
        data=zip_buf.getvalue(),
        file_name="report_by_artist.zip",
        mime="application/zip"
    )


def copy_sheet(src_ws, dst_ws):
    """
    openpyxl에서 시트 전체를 복제할 때, 서식/너비 등을 옮기는 간단한 예시
    """
    from copy import copy
    max_row = src_ws.max_row
    max_col = src_ws.max_column

    for r in range(1, max_row+1):
        dst_ws.row_dimensions[r].height = src_ws.row_dimensions[r].height
        for c in range(1, max_col+1):
            cell_src = src_ws.cell(row=r, column=c)
            cell_dst = dst_ws.cell(row=r, column=c, value=cell_src.value)
            if cell_src.has_style:
                cell_dst.font = copy(cell_src.font)
                cell_dst.border = copy(cell_src.border)
                cell_dst.fill = copy(cell_src.fill)
                cell_dst.number_format = copy(cell_src.number_format)
                cell_dst.protection = copy(cell_src.protection)
                cell_dst.alignment = copy(cell_src.alignment)
    for c in range(1, max_col+1):
        dst_ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = (
            src_ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width
        )


# ========== [4] 핵심 로직: generate_report =============
def generate_report(
    ym: str, 
    report_date: str, 
    check_dict: dict,
    gc: gspread.Client,
    drive_svc,
    sheet_svc,
    progress_bar,
    artist_placeholder
):
    """
    [요약]
    1) input_song cost / input_online revenue 시트에서 해당 ym 데이터를 읽어옴
    2) 아티스트별 매출 및 곡비(전월+당월 발생액, 당월차감 등) 정보를 합산
    3) 구글 스프레드시트 형태의 'output_report_YYYYMM'을 생성하여
       - 각 아티스트별 (1) 세부매출내역 탭, (2) 정산서 탭 생성
       - '정산서' 탭 내 '3. 공제 내역' 칼럼 중 '곡비'를 (전월 잔액 + 당월 발생액)으로 표기
    4) 최종 검증 정보를 check_dict에 누적
    5) 작업 완료 후 out_file_id(생성된 구글시트 ID) 반환
    """

    folder_id = st.secrets["google_service_account_a"]["folder_id"]

    # ------------------- (A) input_song cost -------------------
    try:
        song_cost_sh = gc.open("input_song cost")
    except gspread.exceptions.SpreadsheetNotFound:
        st.error("Google Sheet 'input_song cost'를 찾을 수 없습니다.")
        return ""

    ws_map_sc = {ws.title: ws for ws in song_cost_sh.worksheets()}
    if ym not in ws_map_sc:
        st.error(f"input_song cost에 '{ym}' 탭이 없습니다.")
        return ""
    ws_sc = ws_map_sc[ym]
    data_sc = ws_sc.get_all_values()
    if not data_sc:
        st.error(f"'{ym}' 탭이 비어있습니다.")
        return ""
    header_sc = data_sc[0]
    # 마지막 합계/총계 행은 제외하고 읽는 경우:
    rows_sc = data_sc[1:-1]

    # 이번에 '당월 발생액' 칼럼까지 사용하므로 인덱스 추가
    try:
        idx_sosok  = header_sc.index("소속")  # ← 추가
        idx_artist = header_sc.index("아티스트명")
        idx_rate   = header_sc.index("정산 요율")
        idx_prev   = header_sc.index("전월 잔액")
        idx_curr   = header_sc.index("당월 발생액")
        idx_deduct = header_sc.index("당월 차감액")
        idx_remain = header_sc.index("당월 잔액")
    except ValueError as e:
        st.error(f"[input_song cost] 시트 칼럼 명이 맞는지 확인 필요: {e}")
        return ""

    # 숫자로 변환하는 헬퍼
    def to_num(x: str) -> float:
        if not x:
            return 0.0
        return float(x.replace("%", "").replace(",", ""))


    # 아티스트별 곡비 정보
    #   → '전월 잔액'(prev), '당월 발생액'(curr), '당월 차감'(deduct), '당월 잔액'(remain), '정산요율'(rate)
    #   (실제 작업에서는 나중에 '곡비' = prev + curr)
    artist_cost_dict = {}
    artist_sosok_dict = {}  # ← 새로 추가

    for row in rows_sc:
        artist_name = row[idx_artist].strip()
        if not artist_name:
            continue

        sosok_str = row[idx_sosok].strip().upper()  # "UMAG", "FLUXUS", "UMAG,FLUXUS" 등
        # 만약 쉼표 등으로 여러 소속일 수 있다면 split
        affils = re.split(r'[,&/]', sosok_str)  # ["UMAG","FLUXUS"] 등
        affils = [x.strip() for x in affils if x.strip()]

        cost_data = {
            "정산요율": to_num(row[idx_rate]),
            "전월잔액": to_num(row[idx_prev]),
            "당월발생": to_num(row[idx_curr]),
            "당월차감액": to_num(row[idx_deduct]),
            "당월잔액": to_num(row[idx_remain])
        }
        artist_cost_dict[artist_name] = cost_data

        # artist_sosok_dict에도 넣는다
        artist_sosok_dict[artist_name] = affils  # 예: ["UMAG"], ["FLUXUS"], ["UMAG","FLUXUS"], ...


    # ------------------- (B) input_online revenue (UMAG) -------------
    try:
        revenue_sh = gc.open("input_online revenue_umag_integrated") #UMAG
    except gspread.exceptions.SpreadsheetNotFound:
        st.error("Google Sheet 'input_online revenue_umag_integrated'를 찾을 수 없습니다.")
        return ""

    ws_map_or = {ws.title: ws for ws in revenue_sh.worksheets()}
    if ym not in ws_map_or:
        st.error(f"input_online revenue_umag_integrated에 '{ym}' 탭이 없습니다.")
        return ""
    ws_or = ws_map_or[ym]
    data_or = ws_or.get_all_values()
    if not data_or:
        st.error(f"{ym} 탭이 비어있습니다.")
        return ""

    header_or = data_or[0]
    rows_or = data_or[1:]
    try:
        col_aartist = header_or.index("앨범아티스트")
        col_album   = header_or.index("앨범명")
        col_major   = header_or.index("대분류")
        col_middle  = header_or.index("중분류")
        col_service = header_or.index("서비스명")
        col_revenue = header_or.index("권리사정산금액")
    except ValueError as e:
        st.error(f"[input_online revenue_umag_integrated] 시트 칼럼 명이 맞는지 확인 필요: {e}")
        return ""

    # 아티스트별 매출 정보
    from collections import defaultdict
    artist_revenue_dict = defaultdict(list)
    for row in rows_or:
        a = row[col_aartist].strip()
        if not a:
            continue
        alb = row[col_album]
        maj = row[col_major]
        mid = row[col_middle]
        srv = row[col_service]
        try:
            rv_val = float(row[col_revenue].replace(",", ""))
        except:
            rv_val = 0.0
        artist_revenue_dict[a].append({
            "album": alb,
            "major": maj,
            "middle": mid,
            "service": srv,
            "revenue": rv_val
        })

    # ------------------- (B) input_online revenue (FLUXUS) -------------
    try:
        fluxus_song_sh = gc.open("input_online revenue_fluxus_song")
    except gspread.exceptions.SpreadsheetNotFound:
        st.error("Google Sheet 'input_online revenue_fluxus_song' 없음")
        return ""

    try:
        fluxus_yt_sh = gc.open("input_online revenue_fluxus_yt")
    except gspread.exceptions.SpreadsheetNotFound:
        st.error("Google Sheet 'input_online revenue_fluxus_yt' 없음")
        return ""

    # 여기서 fluxus_song_sh / fluxus_yt_sh 중 ym 탭 존재 여부, get_all_values() 로 읽는 로직, 
    # 그리고 '가수명'(혹은 'ALBIM ARTIST') + '권리사 정산액' 칼럼 파싱 등을 하시면 됩니다.
    # 예시:
    ws_map_fs = {ws.title: ws for ws in fluxus_song_sh.worksheets()}
    if ym not in ws_map_fs:
        st.error(f"fluxus_song '{ym}' 탭 없음")
        return
    ws_fs = ws_map_fs[ym]
    data_fs = ws_fs.get_all_values()

    ws_map_fy = {ws.title: ws for ws in fluxus_yt_sh.worksheets()}
    if ym not in ws_map_fy:
        st.error(f"fluxus_yt '{ym}' 탭 없음")
        return
    ws_fy = ws_map_fy[ym]
    data_fy = ws_fy.get_all_values()


    header_fs = data_fs[0]
    rows_fs = data_fs[1:-1]
    try:
        fs_col_aartist = header_fs.index("가수명")
        fs_col_album   = header_fs.index("앨범명")
        fs_col_country = header_fs.index("서비스 구분")
        fs_col_revenue = header_fs.index("권리사 정산액")
    except ValueError as e:
        st.error(f"[input_online revenue_fluxus_song] 시트 칼럼 명이 맞는지 확인 필요: {e}")
        return ""

    # 아티스트별 매출 정보
    from collections import defaultdict
    fluxus_song_dict = defaultdict(list)
    sum_fs_rv_val = 0.0
    for row in rows_fs:
        a = row[fs_col_aartist].strip()
        if not a:
            continue
        fs_alb = row[fs_col_album]
        fs_ctry = row[fs_col_country]
        try:
            fs_rv_val = float(row[fs_col_revenue].replace(",", ""))
        except:
            fs_rv_val = 0.0
        sum_fs_rv_val += fs_rv_val
        fluxus_song_dict[a].append({
            "album": fs_alb,
            "country": fs_ctry,
            "revenue": fs_rv_val,
            "sum_fs_rv_val": sum_fs_rv_val
        })


    header_fy = data_fy[0]
    rows_fy = data_fy[1:]
    try:
        fy_col_aartist = header_fy.index("ALBIM ARTIST")
        fy_col_album   = header_fy.index("ALBUM TITLE")
        fy_col_title   = header_fy.index("TRACK TITLE")
        fy_col_number  = header_fy.index("TRACK NO.")
        fy_col_id = header_fy.index("TRACK ID")
        fy_col_revenue = header_fy.index("권리사 정산액 \n(KRW)")
    except ValueError as e:
        st.error(f"[input_online revenue_fluxus_yt] 시트 칼럼 명이 맞는지 확인 필요: {e}")
        return ""

    # 아티스트별 매출 정보
    from collections import defaultdict
    fluxus_yt_dict = defaultdict(list)
    sum_fy_rv_val = 0.0
    for row in rows_fy:
        a = row[fy_col_aartist].strip()
        if not a:
            continue
        fy_alb = row[fy_col_album]
        fy_title = row[fy_col_title]
        fy_number = row[fy_col_number]
        fy_id = row[fy_col_id]
        try:
            fy_rv_val = float(row[fy_col_revenue].replace(",", ""))
        except:
            fy_rv_val = 0.0
        sum_fy_rv_val += fy_rv_val
        fluxus_yt_dict[a].append({
            "album": fy_alb,
            "track_title": fy_title,
            "track_number": fy_number,
            "track_id": fy_id,
            "revenue": fy_rv_val,
            "sum_fy_rv_val": sum_fy_rv_val
        })

    
    # for artist, val in fluxus_song_dict.items():
    #    # val이 매출액 float 라면, "앨범명", "서비스명" 등도 함께 append 해야 할 수도 있음
    #    # 예시(단순화):
    #    fluxus_song_dict[artist].append({
    #        "album":  fs_alb,
    #        "country": fs_ctry,
    #        "revenue": fs_rv_val,
    #        "sum_fs_rv_val": sum_fs_rv_val
    #    })

    # for artist, val in fluxus_yt_dict.items():
    #    fluxus_yt_dict[artist].append({
    #        "album": fy_alb,
    #        "track_title": fy_title,
    #        "track_number": fy_number,
    #        "track_id": fy_id,
    #        "revenue": fy_rv_val,
    #        "sum_fy_rv_val": sum_fy_rv_val
    #    })
    
    # ---------------------------------------------------------
    # [추가] check_dict 내부 구조 확인 / 초기화
    # ---------------------------------------------------------
    if "verification_summary" not in check_dict:
        check_dict["verification_summary"] = {
            "total_errors": 0,
            "artist_error_list": []
        }
    if "details_verification" not in check_dict:
        check_dict["details_verification"] = {
            "정산서": [],
            "세부매출": []
        }
    if "details_per_artist" not in check_dict:
        check_dict["details_per_artist"] = {}

    if "fluxus_verification" not in check_dict:
        check_dict["fluxus_verification"] = {
            "정산서": [],
            "세부매출": []
        }

    # 아티스트 목록 검증
    song_artists = [r[idx_artist] for r in rows_sc if r[idx_artist]]
    revenue_artists = [r[col_aartist].strip() for r in rows_or if r[col_aartist].strip()]
    check_dict["song_artists"] = song_artists
    check_dict["revenue_artists"] = revenue_artists

    compare_res = compare_artists(song_artists, revenue_artists)
    check_dict["artist_compare_result"] = compare_res


    # ------------------- (C) 아티스트 목록 ---------------
    all_artists = sorted(set(artist_cost_dict.keys()) | set(artist_revenue_dict.keys()))
    all_artists = [a for a in all_artists if a and a not in ("합계", "총계")]
    st.session_state["all_artists"] = all_artists


    # ------------------- (D) output_report_YYYYMM --------
    out_filename = f"ouput_report_{ym}"
    out_file_id = create_new_spreadsheet(out_filename, folder_id, drive_svc)
    out_sh = gc.open_by_key(out_file_id)
    
    # 기본생성 sheet1 삭제 시도
    try:
        out_sh.del_worksheet(out_sh.worksheet("Sheet1"))
    except:
        pass

    year_val = ym[:4]
    month_val = ym[4:]

    # (UI) 진행률 표시용
    progress_bar.progress(0)
    artist_placeholder.info("아티스트 보고서 생성 중...")

    # 시트 생성(batch)
    needed_titles = []
    for artist in all_artists:
        # 곡비 파일에 아티스트 없으면 스킵!
        if artist not in artist_sosok_dict:
            print(f"[WARN] 곡비에 없는 아티스트 {artist} -> 시트 생성 안 함")
            continue

        affils = artist_sosok_dict[artist]
        for one_sosok in affils:
            if one_sosok == "UMAG":
                needed_titles.append(f"{one_sosok}_{artist}(세부매출내역)")
                needed_titles.append(f"{one_sosok}_{artist}(정산서)")
            elif one_sosok == "FLUXUS":
                needed_titles.append(f"{one_sosok}_{artist}(세부매출내역)")
                needed_titles.append(f"{one_sosok}_{artist}(정산서)")
            else:
                print(f"unknown 소속: {one_sosok}")
    batch_add_sheets(out_file_id, sheet_svc, needed_titles)


    # ===================================================================
    # (E) 아티스트별로 (1) 세부매출내역 탭, (2) 정산서 탭 생성
    # ===================================================================

    all_requests = []  # batchUpdate requests 모음

    for i, artist in enumerate(all_artists):
        if artist not in artist_sosok_dict:
            # 곡비 파일에 없는 아티스트는 스킵(또는 경고 표시)
            print(f"[WARN] 곡비에 없는 아티스트 '{artist}'는 무시합니다.")
            continue
        
        affils = artist_sosok_dict[artist]

        for one_sosok in affils:
            #-------------------------------------------------------------------------
            # UMAG 소속 처리
            #-------------------------------------------------------------------------
            if one_sosok == "UMAG":
                # 진행률
                ratio = (i + 1) / len(all_artists)
                progress_bar.progress(ratio)
                artist_placeholder.info(f"[{i+1}/{len(all_artists)}] '{artist}' 처리 중...")

                # ##################################
                # UMAG 세부매출내역 탭 (batchUpdate 방식)
                # ##################################
                ws_detail = out_sh.worksheet(f"{one_sosok}_{artist}(세부매출내역)")
                details = artist_revenue_dict[artist]
                details_sorted = sorted(details, key=lambda d: album_sort_key(d["album"]))

                detail_matrix = []
                detail_matrix.append(["앨범아티스트","앨범명","대분류","중분류","서비스명","기간","매출 순수익"])

                total_det = 0
                for d in details_sorted:
                    rv = d["revenue"]
                    total_det += rv
                    detail_matrix.append([
                        artist,
                        d["album"],
                        d["major"],
                        d["middle"],
                        d["service"],
                        f"{year_val}년 {month_val}월",
                        to_currency(rv)
                    ])

                # 합계
                detail_matrix.append(["합계","","","","","", to_currency(total_det)])
                row_cursor_detail_end = len(detail_matrix)

                # 시트 업데이트
                ws_detail.update("A1", detail_matrix)
                time.sleep(1)

                # 세부매출내역 탭에 대한 서식/테두리 등 batch 요청
                detail_requests = []
                sheet_id_detail = ws_detail.id

                # (A) 시트 크기(row_cursor_detail_end, 7열)
                detail_requests.append({
                    "updateSheetProperties": {
                        "properties": {
                            "sheetId": sheet_id_detail,
                            "gridProperties": {
                                "rowCount": row_cursor_detail_end,
                                "columnCount": 7
                            }
                        },
                        "fields": "gridProperties(rowCount,columnCount)"
                    }
                })

                # (B) 열너비 설정 (A=0, B=1, ...)
                # 예: A열(0) → 140, B열(1) → 140, E열(4) → 120
                detail_requests.extend([
                    {
                        "updateDimensionProperties": {
                            "range": {
                                "sheetId": sheet_id_detail,
                                "dimension": "COLUMNS",
                                "startIndex": 0, 
                                "endIndex": 1
                            },
                            "properties": {"pixelSize": 140},
                            "fields": "pixelSize"
                        }
                    },
                    {
                        "updateDimensionProperties": {
                            "range": {
                                "sheetId": sheet_id_detail,
                                "dimension": "COLUMNS",
                                "startIndex": 1,
                                "endIndex": 2
                            },
                            "properties": {"pixelSize": 140},
                            "fields": "pixelSize"
                        }
                    },
                    {
                        "updateDimensionProperties": {
                            "range": {
                                "sheetId": sheet_id_detail,
                                "dimension": "COLUMNS",
                                "startIndex": 4,
                                "endIndex": 5
                            },
                            "properties": {"pixelSize": 120},
                            "fields": "pixelSize"
                        }
                    },
                ])

                # (C) 헤더(A1~G1) 포맷
                detail_requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": sheet_id_detail,
                            "startRowIndex": 0,
                            "endRowIndex": 1,
                            "startColumnIndex": 0,
                            "endColumnIndex": 7
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "backgroundColor": {"red":1.0, "green":0.8, "blue":0.0},
                                "horizontalAlignment": "CENTER",
                                "verticalAlignment": "MIDDLE",
                                "textFormat": {
                                    "bold": True,
                                    "foregroundColor": {"red":0,"green":0,"blue":0}
                                }
                            }
                        },
                        "fields": "userEnteredFormat(backgroundColor,horizontalAlignment,verticalAlignment,textFormat)"
                    }
                })

                # (D) 합계행 병합 + 서식
                sum_row_0based = row_cursor_detail_end - 1
                detail_requests.append({
                    "mergeCells": {
                        "range": {
                            "sheetId": sheet_id_detail,
                            "startRowIndex": sum_row_0based,
                            "endRowIndex": sum_row_0based+1,
                            "startColumnIndex": 0,
                            "endColumnIndex": 6
                        },
                        "mergeType": "MERGE_ALL"
                    }
                })
                detail_requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": sheet_id_detail,
                            "startRowIndex": sum_row_0based,
                            "endRowIndex": sum_row_0based+1,
                            "startColumnIndex": 0,
                            "endColumnIndex": 6
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "backgroundColor": {"red":1.0,"green":0.8,"blue":0.0},
                                "horizontalAlignment": "CENTER",
                                "verticalAlignment": "MIDDLE",
                                "textFormat": {"bold": True}
                            }
                        },
                        "fields": "userEnteredFormat(backgroundColor,horizontalAlignment,verticalAlignment,textFormat)"
                    }
                })
                # 합계값(G열)에 오른쪽 정렬
                detail_requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": sheet_id_detail,
                            "startRowIndex": sum_row_0based,
                            "endRowIndex": sum_row_0based+1,
                            "startColumnIndex": 6,
                            "endColumnIndex": 7
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "horizontalAlignment": "RIGHT",
                                "textFormat": {"bold": True}
                            }
                        },
                        "fields": "userEnteredFormat(horizontalAlignment,textFormat)"
                    }
                })
                # 매출 순수익 칼럼 (F열=idx=6) 나머지 행들
                detail_requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": sheet_id_detail,
                            "startRowIndex": 1,
                            "endRowIndex": sum_row_0based,
                            "startColumnIndex": 6,
                            "endColumnIndex": 7
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "horizontalAlignment": "RIGHT"
                            }
                        },
                        "fields": "userEnteredFormat(horizontalAlignment)"
                    }
                })

                # (E) 전체 테두리
                detail_requests.append({
                    "updateBorders": {
                        "range": {
                            "sheetId": sheet_id_detail,
                            "startRowIndex": 0,
                            "endRowIndex": row_cursor_detail_end,
                            "startColumnIndex": 0,
                            "endColumnIndex": 7
                        },
                        "top":    {"style":"SOLID","width":1},
                        "bottom": {"style":"SOLID","width":1},
                        "left":   {"style":"SOLID","width":1},
                        "right":  {"style":"SOLID","width":1},
                        "innerHorizontal": {"style":"SOLID","width":1},
                        "innerVertical":   {"style":"SOLID","width":1}
                    }
                })

                all_requests.extend(detail_requests)

                # (호출 횟수 분할) 1회 batchUpdate 요청이 너무 커지면 나눠서 전송
                if len(all_requests) >= 200:
                    sheet_svc.spreadsheets().batchUpdate(
                        spreadsheetId=out_file_id,
                        body={"requests": all_requests}
                    ).execute()
                    all_requests.clear()
                    time.sleep(1)

                # ##################################
                # UMAG 정산서 탭 (batchUpdate 방식)
                # ##################################
                ws_report = out_sh.worksheet(f"{one_sosok}_{artist}(정산서)")
                ws_report_id = ws_report.id

                # 매출 합
                sum_1 = sum(d["revenue"] for d in details_sorted)  # "음원서비스별" 총합
                # 앨범별 합
                album_sum = defaultdict(float)
                for d in details_sorted:
                    album_sum[d["album"]] += d["revenue"]
                sum_2 = sum(album_sum.values())

                # (A) "곡비" = "전월 잔액 + 당월 발생액" (요청 사항)
                prev_val = artist_cost_dict[artist]["전월잔액"]
                curr_val = artist_cost_dict[artist]["당월발생"]
                # 보고서 '3. 공제 내역'의 '곡비' 칼럼 값 = prev_val + curr_val
                song_cost_for_report = prev_val + curr_val

                # (B) 공제 금액 & 잔액
                deduct_val = artist_cost_dict[artist]["당월차감액"]  # 이미 input_song cost에서 계산된 값
                remain_val = artist_cost_dict[artist]["당월잔액"]   # 동일
                # "공제 적용 후" 매출 = (음원 매출 합) - 공제금액 => sum_2 - deduct_val
                # (단, 요청 사항/업무로직에 따라 정확히 어떻게 적용할지는 케이스별로 맞춤)

                # (C) 정산율 / 최종 정산금액
                rate_val = artist_cost_dict[artist]["정산요율"]
                공제적용후 = sum_2 - deduct_val
                final_amount = 공제적용후 * (rate_val / 100.0)
                

                # --------------------------------------
                # 정산서 테이블(직접 row col 배열 채우기)
                # --------------------------------------
                report_matrix = []
                for _ in range(300):
                    report_matrix.append([""] * 8)

                # 1) 상단 공통정보
                report_matrix[1][6] = report_date   # 보고서 발행일
                report_matrix[3][1] = f"{year_val}년 {month_val}월 판매분"
                report_matrix[5][1] = f"{artist}님 음원 정산 내역서"

                report_matrix[7][0] = "•"
                report_matrix[7][1] = "저희와 함께해 주셔서 정말 감사하고 앞으로도 잘 부탁드리겠습니다!"
                report_matrix[8][0] = "•"
                report_matrix[8][1] = f"{year_val}년 {month_val}월 음원의 수익을 아래와 같이 정산드리오니 참고 부탁드립니다."
                report_matrix[9][0] = "•"
                report_matrix[9][1] = "정산 관련하여 문의사항이 있다면 무엇이든, 언제든 편히 메일 주세요!"
                report_matrix[9][5] = "E-Mail : lucasdh3013@naver.com"

                # -----------------------------------------------------------------
                # 1. 음원 서비스별 정산내역 (세부매출 그대로)
                # -----------------------------------------------------------------
                report_matrix[12][0] = "1."
                report_matrix[12][1] = "음원 서비스별 정산내역"

                header_row_1 = 13
                headers_1 = ["앨범", "대분류", "중분류", "서비스명", "기간", "매출액"]
                for i_h, val_h in enumerate(headers_1):
                    report_matrix[header_row_1][1 + i_h] = val_h

                row_cursor = header_row_1 + 1
                for d in details_sorted:
                    rv = d["revenue"]
                    report_matrix[row_cursor][1] = d["album"]
                    report_matrix[row_cursor][2] = d["major"]
                    report_matrix[row_cursor][3] = d["middle"]
                    report_matrix[row_cursor][4] = d["service"]
                    report_matrix[row_cursor][5] = f"{year_val}년 {month_val}월"
                    report_matrix[row_cursor][6] = to_currency(rv)
                    row_cursor += 1

                row_cursor += 2
                # 합계
                report_matrix[row_cursor-1][1] = "합계"
                report_matrix[row_cursor-1][6] = to_currency(sum_1)
                row_cursor_sum1 = row_cursor
                row_cursor += 1

                # -----------------------------------------------------------------
                # 2. 앨범 별 정산 내역
                # -----------------------------------------------------------------
                report_matrix[row_cursor][0] = "2."
                report_matrix[row_cursor][1] = "앨범 별 정산 내역"
                row_cursor += 1
                row_cursor_album = row_cursor
                report_matrix[row_cursor][1] = "앨범"
                report_matrix[row_cursor][5] = "기간"
                report_matrix[row_cursor][6] = "매출액"
                row_cursor += 1

                for alb in sorted(album_sum.keys(), key=album_sort_key):
                    amt = album_sum[alb]
                    report_matrix[row_cursor][1] = alb
                    report_matrix[row_cursor][5] = f"{year_val}년 {month_val}월"
                    report_matrix[row_cursor][6] = to_currency(amt)
                    row_cursor += 1

                row_cursor += 1
                report_matrix[row_cursor-1][1] = "합계"
                report_matrix[row_cursor-1][6] = to_currency(sum_2)
                row_cursor_sum2 = row_cursor
                row_cursor += 1

                # -----------------------------------------------------------------
                # 3. 공제 내역
                #    (요청사항: '곡비' 칼럼 = (전월 잔액 + 당월 발생액))
                # -----------------------------------------------------------------
                report_matrix[row_cursor][0] = "3."
                report_matrix[row_cursor][1] = "공제 내역"
                row_cursor += 1
                row_cursor_deduction = row_cursor

                report_matrix[row_cursor][1] = "앨범"
                report_matrix[row_cursor][2] = "곡비"
                report_matrix[row_cursor][3] = "공제 금액"
                report_matrix[row_cursor][5] = "공제 후 남은 곡비"
                report_matrix[row_cursor][6] = "공제 적용 금액"
                row_cursor += 1

                # 앨범(들)을 표기만 할지, 혹은 여러줄로 표현할지 등은 업무 규칙에 따라
                alb_list = sorted(album_sum.keys(), key=album_sort_key)
                alb_str = ", ".join(alb_list) if alb_list else "(앨범 없음)"

                report_matrix[row_cursor][1] = alb_str
                # (중요) 여기서 "곡비" = prev_val + curr_val
                report_matrix[row_cursor][2] = to_currency(song_cost_for_report)
                # 공제금액
                report_matrix[row_cursor][3] = to_currency(deduct_val)
                # 공제 후 남은 곡비
                report_matrix[row_cursor][5] = to_currency(remain_val)
                # 공제 적용 금액 (매출 - 공제금액)
                report_matrix[row_cursor][6] = to_currency(sum_2 - deduct_val)
                row_cursor += 2
                row_cursor_sum3 = row_cursor

                # -----------------------------------------------------------------
                # 4. 수익 배분
                # -----------------------------------------------------------------
                report_matrix[row_cursor][0] = "4."
                report_matrix[row_cursor][1] = "수익 배분"
                row_cursor += 1
                row_cursor_rate = row_cursor
                report_matrix[row_cursor][1] = "앨범"
                report_matrix[row_cursor][2] = "항목"
                report_matrix[row_cursor][3] = "적용율"
                report_matrix[row_cursor][6] = "적용 금액"
                row_cursor += 1

                report_matrix[row_cursor][1] = alb_str
                report_matrix[row_cursor][2] = "수익 배분율"
                report_matrix[row_cursor][3] = f"{int(rate_val)}%"
                report_matrix[row_cursor][6] = to_currency(final_amount)
                row_cursor += 1

                report_matrix[row_cursor][1] = "총 정산금액"
                report_matrix[row_cursor][6] = to_currency(final_amount)
                row_cursor_sum4 = row_cursor
                row_cursor += 2

                report_matrix[row_cursor][6] = "* 부가세 별도"
                row_cursor_report_end = row_cursor + 2

                # 시트에 실제 업로드
                ws_report.update("A1", report_matrix)
                time.sleep(1)

                # ------------------------------------
                # (검증) check_dict에 비교결과 반영
                # ------------------------------------
                # (1) 세부매출 vs 정산서
                for d in details_sorted:
                    original_val = d["revenue"]
                    report_val   = d["revenue"]  # 현재는 동일
                    is_match = almost_equal(original_val, report_val)
                    if not is_match:
                        check_dict["verification_summary"]["total_errors"] += 1
                        check_dict["verification_summary"]["artist_error_list"].append(artist)

                    row_report_item = {
                        "아티스트": artist,
                        "구분": "음원서비스별매출",
                        "앨범": d["album"],
                        "서비스명": d["service"],
                        "원본_매출액": original_val,
                        "정산서_매출액": report_val,
                        "match_매출액": is_match,
                    }
                    check_dict["details_verification"]["세부매출"].append(row_report_item)

                # (2) 공제 내역(곡비,공제금액,공제후잔액)
                #   원본(= input_song cost) 값 vs 보고서 값
                #   "곡비"는 (prev + curr), "공제금액"=deduct_val, "남은 곡비"=remain_val
                #   *원본_곡비 = (전월잔액 + 당월발생)
                original_song_cost = artist_cost_dict[artist]["전월잔액"] + artist_cost_dict[artist]["당월발생"]
                is_match_songcost = almost_equal(original_song_cost, song_cost_for_report)
                is_match_deduct   = almost_equal(artist_cost_dict[artist]["당월차감액"], deduct_val)
                is_match_remain   = almost_equal(artist_cost_dict[artist]["당월잔액"], remain_val)

                if not (is_match_songcost and is_match_deduct and is_match_remain):
                    check_dict["verification_summary"]["total_errors"] += 1
                    check_dict["verification_summary"]["artist_error_list"].append(artist)

                row_report_item_3 = {
                    "아티스트": artist,
                    "구분": "공제내역",
                    # 곡비
                    "원본_곡비": original_song_cost,
                    "정산서_곡비": song_cost_for_report,
                    "match_곡비": is_match_songcost,
                    # 공제금액
                    "원본_공제금액": artist_cost_dict[artist]["당월차감액"],
                    "정산서_공제금액": deduct_val,
                    "match_공제금액": is_match_deduct,
                    # 공제후잔액
                    "원본_공제후잔액": artist_cost_dict[artist]["당월잔액"],
                    "정산서_공제후잔액": remain_val,
                    "match_공제후잔액": is_match_remain,
                }
                check_dict["details_verification"]["정산서"].append(row_report_item_3)

                # (3) 4번 수익 배분율
                original_rate = artist_cost_dict[artist]["정산요율"]
                is_rate_match = almost_equal(original_rate, rate_val)
                if not is_rate_match:
                    check_dict["verification_summary"]["total_errors"] += 1
                    check_dict["verification_summary"]["artist_error_list"].append(artist)

                row_report_item_4 = {
                    "아티스트": artist,
                    "구분": "수익배분율",
                    "원본_정산율(%)": original_rate,
                    "정산서_정산율(%)": rate_val,
                    "match_정산율": is_rate_match,
                }
                check_dict["details_verification"]["정산서"].append(row_report_item_4)

                time.sleep(1)   

                # --------------------------------------------------
                # 정산서 탭(디자인/서식) batchUpdate
                # --------------------------------------------------
                report_requests = []

                # (A) 시트 row/col 크기
                report_requests.append({
                    "updateSheetProperties": {
                        "properties": {
                            "sheetId": ws_report_id,
                            "gridProperties": {
                                "rowCount": row_cursor_report_end,
                                "columnCount": 8
                            }
                        },
                        "fields": "gridProperties(rowCount,columnCount)"
                    }
                })

                # (B) 열너비 (A=0 ~ H=7)
                report_requests.extend([
                    {
                        "updateDimensionProperties": {
                            "range": {
                                "sheetId": ws_report_id,
                                "dimension": "COLUMNS",
                                "startIndex": 0,
                                "endIndex": 1
                            },
                            "properties": {"pixelSize": 40},
                            "fields": "pixelSize"
                        }
                    },
                    {
                        "updateDimensionProperties": {
                            "range": {
                                "sheetId": ws_report_id,
                                "dimension": "COLUMNS",
                                "startIndex": 1,
                                "endIndex": 2
                            },
                            "properties": {"pixelSize": 200},
                            "fields": "pixelSize"
                        }
                    },
                    {
                        "updateDimensionProperties": {
                            "range": {
                                "sheetId": ws_report_id,
                                "dimension": "COLUMNS",
                                "startIndex": 2,
                                "endIndex": 3
                            },
                            "properties": {"pixelSize": 130},
                            "fields": "pixelSize"
                        }
                    },
                    {
                        "updateDimensionProperties": {
                            "range": {
                                "sheetId": ws_report_id,
                                "dimension": "COLUMNS",
                                "startIndex": 3,
                                "endIndex": 4
                            },
                            "properties": {"pixelSize": 120},
                            "fields": "pixelSize"
                        }
                    },
                    {
                        "updateDimensionProperties": {
                            "range": {
                                "sheetId": ws_report_id,
                                "dimension": "COLUMNS",
                                "startIndex": 4,
                                "endIndex": 5
                            },
                            "properties": {"pixelSize": 130},
                            "fields": "pixelSize"
                        }
                    },
                    {
                        "updateDimensionProperties": {
                            "range": {
                                "sheetId": ws_report_id,
                                "dimension": "COLUMNS",
                                "startIndex": 5,
                                "endIndex": 6
                            },
                            "properties": {"pixelSize": 130},
                            "fields": "pixelSize"
                        }
                    },
                    {
                        "updateDimensionProperties": {
                            "range": {
                                "sheetId": ws_report_id,
                                "dimension": "COLUMNS",
                                "startIndex": 6,
                                "endIndex": 7
                            },
                            "properties": {"pixelSize": 130},
                            "fields": "pixelSize"
                        }
                    },
                    {
                        "updateDimensionProperties": {
                            "range": {
                                "sheetId": ws_report_id,
                                "dimension": "COLUMNS",
                                "startIndex": 7,
                                "endIndex": 8
                            },
                            "properties": {"pixelSize": 40},
                            "fields": "pixelSize"
                        }
                    },
                ])

                # (C) 특정행 높이
                report_requests.extend([
                    {
                        "updateDimensionProperties": {
                            "range": {
                                "sheetId": ws_report_id,
                                "dimension": "ROWS",
                                "startIndex": 3,
                                "endIndex": 4
                            },
                            "properties": {"pixelSize": 30},
                            "fields": "pixelSize"
                        }
                    },
                    {
                        "updateDimensionProperties": {
                            "range": {
                                "sheetId": ws_report_id,
                                "dimension": "ROWS",
                                "startIndex": 5,
                                "endIndex": 6
                            },
                            "properties": {"pixelSize": 30},
                            "fields": "pixelSize"
                        }
                    },
                ])

                # (D) 상단 고정 항목(발행 날짜, H2: row=1, col=6)
                report_requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": ws_report_id,
                            "startRowIndex": 1,
                            "endRowIndex": 2,
                            "startColumnIndex": 6,
                            "endColumnIndex": 7
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "horizontalAlignment": "RIGHT",
                                "verticalAlignment": "MIDDLE",
                                "textFormat": {
                                    "fontFamily": "Malgun Gothic",
                                    "fontSize": 10,
                                    "bold": False
                                }
                            }
                        },
                        "fields": "userEnteredFormat(horizontalAlignment,verticalAlignment,textFormat)"
                    }
                })        

                # (E) 상단 고정 항목(판매분, B4:E4)
                report_requests.append({
                    "mergeCells": {
                        "range": {
                            "sheetId": ws_report_id,
                            "startRowIndex": 3,  # (4-1)
                            "endRowIndex": 4,
                            "startColumnIndex": 1,  # (B=1)
                            "endColumnIndex": 5     # (E=4 => endIndex=5)
                        },
                        "mergeType": "MERGE_ALL"
                    }
                })
                report_requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": ws_report_id,
                            "startRowIndex": 3,
                            "endRowIndex": 4,
                            "startColumnIndex": 1,
                            "endColumnIndex": 5
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "horizontalAlignment": "LEFT",
                                "verticalAlignment": "MIDDLE",
                                "textFormat": {
                                    "fontFamily": "Malgun Gothic",
                                    "fontSize": 15,
                                    "bold": True
                                }
                            }
                        },
                        "fields": "userEnteredFormat(horizontalAlignment,verticalAlignment,textFormat)"
                    }
                })

                # (F) 상단 고정 항목(아티스트 정산내역서, B6:G6)
                report_requests.append({
                    "mergeCells": {
                        "range": {
                            "sheetId": ws_report_id,
                            "startRowIndex": 5,
                            "endRowIndex": 6,
                            "startColumnIndex": 1,
                            "endColumnIndex": 7
                        },
                        "mergeType": "MERGE_ALL"
                    }
                })
                report_requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": ws_report_id,
                            "startRowIndex": 5,
                            "endRowIndex": 6,
                            "startColumnIndex": 1,
                            "endColumnIndex": 7
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "backgroundColor": {"red": 0.896, "green": 0.988, "blue": 1},
                                "horizontalAlignment": "CENTER",
                                "verticalAlignment": "MIDDLE",
                                "textFormat": {
                                    "fontFamily": "Malgun Gothic",
                                    "fontSize": 15,
                                    "bold": True
                                }
                            }
                        },
                        "fields": "userEnteredFormat(backgroundColor,horizontalAlignment,verticalAlignment,textFormat)"
                    }
                })

                # (G) 상단 고정 항목(안내문, B8:E8~B10:E10)
                #8행
                report_requests.append({
                    "mergeCells": {
                        "range": {
                            "sheetId": ws_report_id,
                            "startRowIndex": 7,  # (4-1)
                            "endRowIndex": 8,
                            "startColumnIndex": 1,  # (B=1)
                            "endColumnIndex": 5     # (E=4 => endIndex=5)
                        },
                        "mergeType": "MERGE_ALL"
                    }
                })
                report_requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": ws_report_id,
                            "startRowIndex": 7,  # (4-1)
                            "endRowIndex": 8,
                            "startColumnIndex": 1,  # (B=1)
                            "endColumnIndex": 5     # (E=4 => endIndex=5)
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "horizontalAlignment": "LEFT",
                                "verticalAlignment": "MIDDLE",
                                "textFormat": {
                                    "bold": False
                                }
                            }
                        },
                        "fields": "userEnteredFormat(horizontalAlignment,verticalAlignment,textFormat)"
                    }
                })
                #9행
                report_requests.append({
                    "mergeCells": {
                        "range": {
                            "sheetId": ws_report_id,
                            "startRowIndex": 8,  
                            "endRowIndex": 9,
                            "startColumnIndex": 1,
                            "endColumnIndex": 5 
                        },
                        "mergeType": "MERGE_ALL"
                    }
                })
                report_requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": ws_report_id,
                            "startRowIndex": 8,  
                            "endRowIndex": 9,
                            "startColumnIndex": 1,
                            "endColumnIndex": 5 
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "horizontalAlignment": "LEFT",
                                "verticalAlignment": "MIDDLE",
                                "textFormat": {
                                "bold": False
                                }
                            }
                        },
                        "fields": "userEnteredFormat(horizontalAlignment,verticalAlignment,textFormat)"
                    }
                })
                #10행
                report_requests.append({
                    "mergeCells": {
                        "range": {
                            "sheetId": ws_report_id,
                            "startRowIndex": 9,  
                            "endRowIndex": 10,
                            "startColumnIndex": 1,
                            "endColumnIndex": 5 
                        },
                        "mergeType": "MERGE_ALL"
                    }
                })
                report_requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": ws_report_id,
                            "startRowIndex": 9,  
                            "endRowIndex": 10,
                            "startColumnIndex": 1,
                            "endColumnIndex": 5 
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "horizontalAlignment": "LEFT",
                                "verticalAlignment": "MIDDLE",
                                "textFormat": {
                                    "bold": False
                                }
                            }
                        },
                        "fields": "userEnteredFormat(horizontalAlignment,verticalAlignment,textFormat)"
                    }
                })
                # 10행 (E-Mail 칸)
                report_requests.append({
                    "mergeCells": {
                        "range": {
                            "sheetId": ws_report_id,
                            "startRowIndex": 9,  
                            "endRowIndex": 10,
                            "startColumnIndex": 5,
                            "endColumnIndex": 7 
                        },
                        "mergeType": "MERGE_ALL"
                    }
                })
                report_requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": ws_report_id,
                            "startRowIndex": 9,
                            "endRowIndex": 10,
                            "startColumnIndex": 5,
                            "endColumnIndex": 7
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "horizontalAlignment": "RIGHT",
                                "verticalAlignment": "MIDDLE",
                                "textFormat": {
                                    "foregroundColor": {"red": 0.29, "green": 0.53, "blue": 0.91},
                                    "fontFamily": "Malgun Gothic",
                                    "fontSize": 10,
                                    "bold": True
                                }
                            }
                        },
                        "fields": "userEnteredFormat(horizontalAlignment,verticalAlignment,textFormat)"
                    }
                })
                
                # (H) 1열 정렬 (번호 영역)
                report_requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": ws_report_id,
                            "startRowIndex": 1,
                            "endRowIndex": row_cursor_rate+1,
                            "startColumnIndex": 0,
                            "endColumnIndex": 1
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "horizontalAlignment": "CENTER",
                                "verticalAlignment": "MIDDLE",
                                "textFormat": {
                                    "fontFamily": "Malgun Gothic",
                                    "fontSize": 10,
                                    "bold": False
                                }
                            }
                        },
                        "fields": "userEnteredFormat(horizontalAlignment,verticalAlignment,textFormat)"
                    }
                })

                # (I) 하단 고정 항목(부가세, G)
                report_requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": ws_report_id,
                            "startRowIndex": row_cursor_report_end-2,
                            "endRowIndex": row_cursor_report_end,
                            "startColumnIndex": 6,
                            "endColumnIndex": 7
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "horizontalAlignment": "RIGHT",
                                "verticalAlignment": "MIDDLE",
                                "textFormat": {
                                    "fontFamily": "Malgun Gothic",
                                    "fontSize": 10,
                                    "bold": False
                                }
                            }
                        },
                        "fields": "userEnteredFormat(horizontalAlignment,verticalAlignment,textFormat)"
                    }
                }) 
            

                # (J-1) "음원 서비스별 정산내역" 표 타이틀
                report_requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": ws_report_id,
                            "startRowIndex": 12,  # (4-1)
                            "endRowIndex": 13,
                            "startColumnIndex": 1,  # (B=1)
                            "endColumnIndex": 2     # (E=4 => endIndex=5)
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "horizontalAlignment": "LEFT",
                                "verticalAlignment": "MIDDLE",
                                "textFormat": {
                                    "bold": True
                                }
                            }
                        },
                        "fields": "userEnteredFormat(horizontalAlignment,verticalAlignment,textFormat)"
                    }
                })
                # (J-2) "음원 서비스별 정산내역" 표 헤더 (Row=13)
                report_requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": ws_report_id,
                            "startRowIndex": 13,
                            "endRowIndex": 14,
                            "startColumnIndex": 1,
                            "endColumnIndex": 7
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "backgroundColor": {"red": 0.3, "green": 0.82, "blue": 0.88},
                                "horizontalAlignment": "CENTER",
                                "verticalAlignment": "MIDDLE",
                                "textFormat": {
                                    "fontFamily": "Malgun Gothic",
                                    "fontSize": 10,
                                    "bold": True
                                }
                            }
                        },
                        "fields": "userEnteredFormat(backgroundColor,horizontalAlignment,verticalAlignment,textFormat)"
                    }
                })
                # (J-3) 합계행 전 병합
                report_requests.append({
                    "mergeCells": {
                        "range": {
                            "sheetId": ws_report_id,
                            "startRowIndex": row_cursor_sum1-2,
                            "endRowIndex": row_cursor_sum1-1,
                            "startColumnIndex": 1,
                            "endColumnIndex": 7
                        },
                        "mergeType": "MERGE_ALL"
                    }
                })
                # (J-4) 합계행 병합
                report_requests.append({
                    "mergeCells": {
                        "range": {
                            "sheetId": ws_report_id,
                            "startRowIndex": row_cursor_sum1-1,
                            "endRowIndex": row_cursor_sum1,
                            "startColumnIndex": 1,
                            "endColumnIndex": 6
                        },
                        "mergeType": "MERGE_ALL"
                    }
                })
                report_requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": ws_report_id,
                            "startRowIndex": row_cursor_sum1-1,
                            "endRowIndex": row_cursor_sum1,
                            "startColumnIndex": 1,
                            "endColumnIndex": 6
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "backgroundColor": {"red": 0.896, "green": 0.988, "blue": 1},
                                "horizontalAlignment": "CENTER",
                                "verticalAlignment": "MIDDLE",
                                "textFormat": {
                                    "fontFamily": "Malgun Gothic",
                                    "fontSize": 10,
                                    "bold": True
                                }
                            }
                        },
                        "fields": "userEnteredFormat(backgroundColor,horizontalAlignment,verticalAlignment,textFormat)"
                    }
                })
                report_requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": ws_report_id,
                            "startRowIndex": row_cursor_sum1-1,
                            "endRowIndex": row_cursor_sum1,
                            "startColumnIndex": 6,
                            "endColumnIndex": 7
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "backgroundColor": {"red": 0.896, "green": 0.988, "blue": 1},
                                "horizontalAlignment": "CENTER",
                                "verticalAlignment": "MIDDLE",
                                "textFormat": {
                                    "fontFamily": "Malgun Gothic",
                                    "fontSize": 10,
                                    "bold": True
                                }
                            }
                        },
                        "fields": "userEnteredFormat(backgroundColor,horizontalAlignment,verticalAlignment,textFormat)"
                    }
                })
                # (J-5) 표에 Banding (줄무늬 효과)
                banding_start_row = 14
                banding_end_row = row_cursor_sum1 - 2
                banding_start_col = 1
                banding_end_col = 7
                if banding_end_row > banding_start_row:  # 유효범위 체크
                    report_requests.append({
                        "addBanding": {
                            "bandedRange": {
                                "range": {
                                    "sheetId": ws_report_id,
                                    "startRowIndex": banding_start_row,
                                    "endRowIndex": banding_end_row,
                                    "startColumnIndex": banding_start_col,
                                    "endColumnIndex": banding_end_col
                                },
                                "rowProperties": {
                                    "firstBandColor": {
                                        "red": 1.0, "green": 1.0, "blue": 1.0
                                    },
                                    "secondBandColor": {
                                        "red": 0.896, "green": 0.988, "blue": 1
                                    }
                                },
                                
                            }
                        }
                    })
                    report_requests.append({
                        "repeatCell": {
                            "range": {
                                "sheetId": ws_report_id,
                                "startRowIndex": banding_start_row,
                                "endRowIndex": banding_end_row,
                                "startColumnIndex": banding_start_col,
                                "endColumnIndex": banding_end_col
                            },
                            "cell": {
                                "userEnteredFormat": {
                                    "horizontalAlignment": "CENTER",
                                    "verticalAlignment": "MIDDLE",
                                    "textFormat": {
                                        "fontFamily": "Malgun Gothic",
                                        "fontSize": 10,
                                        "bold": False
                                    }
                                }
                            },
                            "fields": "userEnteredFormat(horizontalAlignment,verticalAlignment,textFormat)"
                        }
                    })


                # (K-1) 앨범별 정산내역 타이틀
                report_requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": ws_report_id,
                            "startRowIndex": row_cursor_album-1,
                            "endRowIndex": row_cursor_album,
                            "startColumnIndex": 1, 
                            "endColumnIndex": 2    
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "horizontalAlignment": "LEFT",
                                "verticalAlignment": "MIDDLE",
                                "textFormat": {
                                    "bold": True
                                }
                            }
                        },
                        "fields": "userEnteredFormat(horizontalAlignment,verticalAlignment,textFormat)"
                    }
                })
                # (K-2) 앨범별 정산내역 헤더
                report_requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": ws_report_id,
                            "startRowIndex": row_cursor_album,
                            "endRowIndex": row_cursor_album+1,
                            "startColumnIndex": 1,
                            "endColumnIndex": 7
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "backgroundColor": {"red": 0.3, "green": 0.82, "blue": 0.88},
                                "horizontalAlignment": "CENTER",
                                "verticalAlignment": "MIDDLE",
                                "textFormat": {
                                    "fontFamily": "Malgun Gothic",
                                    "fontSize": 10,
                                    "bold": True
                                }
                            }
                        },
                        "fields": "userEnteredFormat(backgroundColor,horizontalAlignment,verticalAlignment,textFormat)"
                    }
                })
                # (K-3) 앨범별 정산내역 표 본문
                report_requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": ws_report_id,
                            "startRowIndex": row_cursor_album+1,
                            "endRowIndex": row_cursor_sum2-1,
                            "startColumnIndex": 1,
                            "endColumnIndex": 7
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "horizontalAlignment": "CENTER",
                                "verticalAlignment": "MIDDLE",
                                "textFormat": {
                                    "fontFamily": "Malgun Gothic",
                                    "fontSize": 10,
                                    "bold": False
                                }
                            }
                        },
                        "fields": "userEnteredFormat(horizontalAlignment,verticalAlignment,textFormat)"
                    }
                })
                # (K-4) 앨범별 정산내역 합계행
                report_requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": ws_report_id,
                            "startRowIndex": row_cursor_sum2-1,
                            "endRowIndex": row_cursor_sum2,
                            "startColumnIndex": 1,
                            "endColumnIndex": 7
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "horizontalAlignment": "CENTER",
                                "verticalAlignment": "MIDDLE",
                                "textFormat": {
                                    "fontFamily": "Malgun Gothic",
                                    "fontSize": 10,
                                    "bold": True
                                }
                            }
                        },
                        "fields": "userEnteredFormat(horizontalAlignment,verticalAlignment,textFormat)"
                    }
                })
                # (K-5) 합계행 병합
                report_requests.append({
                    "mergeCells": {
                        "range": {
                            "sheetId": ws_report_id,
                            "startRowIndex": row_cursor_sum2-1,
                            "endRowIndex": row_cursor_sum2,
                            "startColumnIndex": 1,
                            "endColumnIndex": 6
                        },
                        "mergeType": "MERGE_ALL"
                    }
                })
                report_requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": ws_report_id,
                            "startRowIndex": row_cursor_sum2-1,
                            "endRowIndex": row_cursor_sum2,
                            "startColumnIndex": 1,
                            "endColumnIndex": 6
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "backgroundColor": {"red": 0.896, "green": 0.988, "blue": 1},
                                "horizontalAlignment": "CENTER",
                                "verticalAlignment": "MIDDLE",
                                "textFormat": {
                                    "fontFamily": "Malgun Gothic",
                                    "fontSize": 10,
                                    "bold": True
                                }
                            }
                        },
                        "fields": "userEnteredFormat(backgroundColor,horizontalAlignment,verticalAlignment,textFormat)"
                    }
                })
                report_requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": ws_report_id,
                            "startRowIndex": row_cursor_sum2-1,
                            "endRowIndex": row_cursor_sum2,
                            "startColumnIndex": 6,
                            "endColumnIndex": 7
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "backgroundColor": {"red": 0.896, "green": 0.988, "blue": 1},
                                "horizontalAlignment": "CENTER",
                                "verticalAlignment": "MIDDLE",
                                "textFormat": {
                                    "fontFamily": "Malgun Gothic",
                                    "fontSize": 10,
                                    "bold": True
                                }
                            }
                        },
                        "fields": "userEnteredFormat(backgroundColor,horizontalAlignment,verticalAlignment,textFormat)"
                    }
                })


                # (L-1) 공제 내역 타이틀
                report_requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": ws_report_id,
                            "startRowIndex": row_cursor_deduction-1,  # (4-1)
                            "endRowIndex": row_cursor_deduction,
                            "startColumnIndex": 1,  # (B=1)
                            "endColumnIndex": 2     # (E=4 => endIndex=5)
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "horizontalAlignment": "LEFT",
                                "verticalAlignment": "MIDDLE",
                                "textFormat": {
                                    "bold": True
                                }
                            }
                        },
                        "fields": "userEnteredFormat(horizontalAlignment,verticalAlignment,textFormat)"
                    }
                })
                # (L-2) 공제 내역 헤더
                report_requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": ws_report_id,
                            "startRowIndex": row_cursor_deduction,
                            "endRowIndex": row_cursor_deduction+1,
                            "startColumnIndex": 1,
                            "endColumnIndex": 7
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "backgroundColor": {"red": 0.3, "green": 0.82, "blue": 0.88},
                                "horizontalAlignment": "CENTER",
                                "verticalAlignment": "MIDDLE",
                                "textFormat": {
                                    "fontFamily": "Malgun Gothic",
                                    "fontSize": 10,
                                    "bold": True
                                }
                            }
                        },
                        "fields": "userEnteredFormat(backgroundColor,horizontalAlignment,verticalAlignment,textFormat)"
                    }
                })
                # (L-3) 공제 내역 표 본문 (데이터부분)
                report_requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": ws_report_id,
                            "startRowIndex": row_cursor_deduction+1,
                            "endRowIndex": row_cursor_deduction+2,
                            "startColumnIndex": 1,
                            "endColumnIndex": 6
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "horizontalAlignment": "CENTER",
                                "verticalAlignment": "MIDDLE",
                                "textFormat": {
                                    "fontFamily": "Malgun Gothic",
                                    "fontSize": 10,
                                    "bold": False
                                }
                            }
                        },
                        "fields": "userEnteredFormat(horizontalAlignment,verticalAlignment,textFormat)"
                    }
                })
                # (L-4) 공제 내역 표 본문 (합계 부분)
                report_requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": ws_report_id,
                            "startRowIndex": row_cursor_deduction+1,
                            "endRowIndex": row_cursor_deduction+2,
                            "startColumnIndex": 6,
                            "endColumnIndex": 7
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "horizontalAlignment": "CENTER",
                                "verticalAlignment": "MIDDLE",
                                "textFormat": {
                                    "fontFamily": "Malgun Gothic",
                                    "fontSize": 10,
                                    "bold": True
                                }
                            }
                        },
                        "fields": "userEnteredFormat(horizontalAlignment,verticalAlignment,textFormat)"
                    }
                })


                # (M-1) 수익 배분 타이틀
                report_requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": ws_report_id,
                            "startRowIndex": row_cursor_rate-1,
                            "endRowIndex": row_cursor_rate,
                            "startColumnIndex": 1,  
                            "endColumnIndex": 2    
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "horizontalAlignment": "LEFT",
                                "verticalAlignment": "MIDDLE",
                                "textFormat": {
                                    "bold": True
                                }
                            }
                        },
                        "fields": "userEnteredFormat(horizontalAlignment,verticalAlignment,textFormat)"
                    }
                })
                # (M-2) 수익 배분 헤더
                report_requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": ws_report_id,
                            "startRowIndex": row_cursor_rate,
                            "endRowIndex": row_cursor_rate+1,
                            "startColumnIndex": 1,
                            "endColumnIndex": 7
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "backgroundColor": {"red": 0.3, "green": 0.82, "blue": 0.88},
                                "horizontalAlignment": "CENTER",
                                "verticalAlignment": "MIDDLE",
                                "textFormat": {
                                    "fontFamily": "Malgun Gothic",
                                    "fontSize": 10,
                                    "bold": True
                                }
                            }
                        },
                        "fields": "userEnteredFormat(backgroundColor,horizontalAlignment,verticalAlignment,textFormat)"
                    }
                })
                # (M-3) 수익 배분 표 본문 
                report_requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": ws_report_id,
                            "startRowIndex": row_cursor_rate+1,
                            "endRowIndex": row_cursor_rate+2,
                            "startColumnIndex": 1,
                            "endColumnIndex": 7
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "horizontalAlignment": "CENTER",
                                "verticalAlignment": "MIDDLE",
                                "textFormat": {
                                    "fontFamily": "Malgun Gothic",
                                    "fontSize": 10,
                                    "bold": False
                                }
                            }
                        },
                        "fields": "userEnteredFormat(horizontalAlignment,verticalAlignment,textFormat)"
                    }
                })
                # (M-4) 수익 배분 표 합계행 병합
                report_requests.append({
                    "mergeCells": {
                        "range": {
                            "sheetId": ws_report_id,
                            "startRowIndex": row_cursor_sum4,
                            "endRowIndex": row_cursor_sum4+1,
                            "startColumnIndex": 1,
                            "endColumnIndex": 6
                        },
                        "mergeType": "MERGE_ALL"
                    }
                })
                report_requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": ws_report_id,
                            "startRowIndex": row_cursor_sum4,
                            "endRowIndex": row_cursor_sum4+1,
                            "startColumnIndex": 1,
                            "endColumnIndex": 6
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "backgroundColor": {"red": 0.896, "green": 0.988, "blue": 1},
                                "horizontalAlignment": "CENTER",
                                "verticalAlignment": "MIDDLE",
                                "textFormat": {
                                    "fontFamily": "Malgun Gothic",
                                    "fontSize": 10,
                                    "bold": True
                                }
                            }
                        },
                        "fields": "userEnteredFormat(backgroundColor,horizontalAlignment,verticalAlignment,textFormat)"
                    }
                })
                report_requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": ws_report_id,
                            "startRowIndex": row_cursor_sum4,
                            "endRowIndex": row_cursor_sum4+1,
                            "startColumnIndex": 6,
                            "endColumnIndex": 7
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "backgroundColor": {"red": 0.896, "green": 0.988, "blue": 1},
                                "horizontalAlignment": "CENTER",
                                "verticalAlignment": "MIDDLE",
                                "textFormat": {
                                    "fontFamily": "Malgun Gothic",
                                    "fontSize": 10,
                                    "bold": True
                                }
                            }
                        },
                        "fields": "userEnteredFormat(backgroundColor,horizontalAlignment,verticalAlignment,textFormat)"
                    }
                })


                # (N) 전체 테두리 화이트
                report_requests.append({
                    "updateBorders": {
                        "range": {
                            "sheetId": ws_report_id,
                            "startRowIndex": 0,
                            "endRowIndex": row_cursor_report_end,
                            "startColumnIndex": 0,
                            "endColumnIndex": 8
                        },
                        "top":    {"style": "SOLID","width":1, "color":{"red":1,"green":1,"blue":1}},
                        "bottom": {"style": "SOLID","width":1, "color":{"red":1,"green":1,"blue":1}},
                        "left":   {"style": "SOLID","width":1, "color":{"red":1,"green":1,"blue":1}},
                        "right":  {"style": "SOLID","width":1, "color":{"red":1,"green":1,"blue":1}},
                        "innerHorizontal": {"style":"SOLID","width":1,"color":{"red":1,"green":1,"blue":1}},
                        "innerVertical":   {"style":"SOLID","width":1,"color":{"red":1,"green":1,"blue":1}}
                    }
                })
                

                # (O) 표 부분 점선 
                def add_dotted_borders(r1, r2, c1, c2):
                    """바깥+안쪽 모두 DOTTED"""
                    report_requests.append({
                        "updateBorders": {
                            "range": {
                                "sheetId": ws_report_id,
                                "startRowIndex": r1,
                                "endRowIndex": r2,
                                "startColumnIndex": c1,
                                "endColumnIndex": c2
                            },
                            "top":    {"style": "DOTTED", "width": 1, "color":{"red":0,"green":0,"blue":0}},
                            "bottom": {"style": "DOTTED", "width": 1, "color":{"red":0,"green":0,"blue":0}},
                            "left":   {"style": "DOTTED", "width": 1, "color":{"red":0,"green":0,"blue":0}},
                            "right":  {"style": "DOTTED", "width": 1, "color":{"red":0,"green":0,"blue":0}},
                            "innerHorizontal": {"style": "DOTTED", "width": 1, "color":{"red":0,"green":0,"blue":0}},
                            "innerVertical":   {"style": "DOTTED", "width": 1, "color":{"red":0,"green":0,"blue":0}}
                        }
                    })
                # 1번 섹션 A14:G30 => row=13..30, col=0..7
                add_dotted_borders(13, row_cursor_sum1, 1, 7)
                # 2번 섹션 
                add_dotted_borders(row_cursor_album, row_cursor_sum2, 1, 7)
                # 3번 섹션 
                add_dotted_borders(row_cursor_deduction, row_cursor_sum3-1, 1, 7)
                # 4번 섹션 
                add_dotted_borders(row_cursor_rate, row_cursor_sum4+1, 1, 7)
                

                # (P) 시트 외곽 검정 SOLID 
                report_requests.append({
                    "updateBorders": {
                        "range": {
                            "sheetId": ws_report_id,
                            "startRowIndex": 0,
                            "endRowIndex": row_cursor_report_end,
                            "startColumnIndex": 0,
                            "endColumnIndex": 8
                        },
                        "top":    {"style": "SOLID","width":1, "color":{"red":0,"green":0,"blue":0}},
                        "bottom": {"style": "SOLID","width":1, "color":{"red":0,"green":0,"blue":0}},
                        "left":   {"style": "SOLID","width":1, "color":{"red":0,"green":0,"blue":0}},
                        "right":  {"style": "SOLID","width":1, "color":{"red":0,"green":0,"blue":0}}
                        # innerHorizontal, innerVertical는 생략 => 기존 값 유지
                    }
                })
                
                all_requests.extend(report_requests)


                # batchUpdate 분할 전송
                if len(all_requests) >= 200:
                    sheet_svc.spreadsheets().batchUpdate(
                        spreadsheetId=out_file_id,
                        body={"requests": all_requests}
                    ).execute()
                    all_requests.clear()
                    time.sleep(1)

            #-------------------------------------------------------------------------
            # FLUXUS 소속 처리
            #-------------------------------------------------------------------------
            elif one_sosok == "FLUXUS":
                # 진행률
                ratio = (i + 1) / len(all_artists)
                progress_bar.progress(ratio)
                artist_placeholder.info(f"[{i+1}/{len(all_artists)}] '{artist}' 처리 중...")

                # ##########################
                # (1) FLUXUS 세부매출내역 탭
                # ##########################
                ws_fluxus_detail = out_sh.worksheet(f"{one_sosok}_{artist}(세부매출내역)")
                fluxus_yt_details = fluxus_yt_dict[artist]
                fluxus_fs_details = fluxus_song_dict[artist]
                fluxus_yt_details_sorted = sorted(fluxus_yt_details, key=lambda d: album_sort_key(d["album"]))
                fluxus_fs_details_sorted = sorted(fluxus_fs_details, key=lambda d: album_sort_key(d["album"]))

                fluxus_detail_matrix = []
                fluxus_detail_matrix.append(["앨범아티스트","앨범명","트랙 No.","트랙명","트랙 ID","기간","매출 순수익"])

                total_det = 0
                for d in fluxus_yt_details_sorted:
                    fy_rv_val = d["revenue"]
                    total_det += fy_rv_val
                    fluxus_detail_matrix.append([
                        artist,
                        d["album"],
                        d["track_number"],
                        d["track_title"],
                        d["track_id"],
                        f"{year_val}년 {month_val}월",
                        to_currency(fy_rv_val)
                    ])

                # 합계
                fluxus_detail_matrix.append(["합계","","","","","", to_currency(total_det)])
                row_cursor_fluxus_detail_end = len(fluxus_detail_matrix)

                # 시트 업데이트
                ws_fluxus_detail.update("A1", fluxus_detail_matrix)
                time.sleep(1)

                # 세부매출내역 탭에 대한 서식/테두리 등 batch 요청
                fluxus_detail_requests = []
                sheet_id_fluxus_detail = ws_fluxus_detail.id

                # (A) 시트 크기(row_cursor_detail_end, 7열)
                fluxus_detail_requests.append({
                    "updateSheetProperties": {
                        "properties": {
                            "sheetId": sheet_id_fluxus_detail,
                            "gridProperties": {
                                "rowCount": row_cursor_fluxus_detail_end,
                                "columnCount": 7
                            }
                        },
                        "fields": "gridProperties(rowCount,columnCount)"
                    }
                })

                # (B) 열너비 설정 (A=0, B=1, ...)
                # 예: A열(0) → 140, B열(1) → 140, E열(4) → 120
                fluxus_detail_requests.extend([
                    {
                        "updateDimensionProperties": {
                            "range": {
                                "sheetId": sheet_id_fluxus_detail,
                                "dimension": "COLUMNS",
                                "startIndex": 0, 
                                "endIndex": 1
                            },
                            "properties": {"pixelSize": 140},
                            "fields": "pixelSize"
                        }
                    },
                    {
                        "updateDimensionProperties": {
                            "range": {
                                "sheetId": sheet_id_fluxus_detail,
                                "dimension": "COLUMNS",
                                "startIndex": 1,
                                "endIndex": 2
                            },
                            "properties": {"pixelSize": 160},
                            "fields": "pixelSize"
                        }
                    },
                    {
                        "updateDimensionProperties": {
                            "range": {
                                "sheetId": sheet_id_fluxus_detail,
                                "dimension": "COLUMNS",
                                "startIndex": 2,
                                "endIndex": 3
                            },
                            "properties": {"pixelSize": 100},
                            "fields": "pixelSize"
                        }
                    },
                                        {
                        "updateDimensionProperties": {
                            "range": {
                                "sheetId": sheet_id_fluxus_detail,
                                "dimension": "COLUMNS",
                                "startIndex": 3,
                                "endIndex": 4
                            },
                            "properties": {"pixelSize": 160},
                            "fields": "pixelSize"
                        }
                    },
                                        {
                        "updateDimensionProperties": {
                            "range": {
                                "sheetId": sheet_id_fluxus_detail,
                                "dimension": "COLUMNS",
                                "startIndex": 4,
                                "endIndex": 5
                            },
                            "properties": {"pixelSize": 100},
                            "fields": "pixelSize"
                        }
                    },
                                        {
                        "updateDimensionProperties": {
                            "range": {
                                "sheetId": sheet_id_fluxus_detail,
                                "dimension": "COLUMNS",
                                "startIndex": 5,
                                "endIndex": 6
                            },
                            "properties": {"pixelSize": 120},
                            "fields": "pixelSize"
                        }
                    },
                                        {
                        "updateDimensionProperties": {
                            "range": {
                                "sheetId": sheet_id_fluxus_detail,
                                "dimension": "COLUMNS",
                                "startIndex": 6,
                                "endIndex": 7
                            },
                            "properties": {"pixelSize": 140},
                            "fields": "pixelSize"
                        }
                    }
                ])

                # (C) 헤더(A1~G1) 포맷
                fluxus_detail_requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": sheet_id_fluxus_detail,
                            "startRowIndex": 0,
                            "endRowIndex": 1,
                            "startColumnIndex": 0,
                            "endColumnIndex": 7
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "backgroundColor": {"red": 0.3, "green": 0.82, "blue": 0.88},
                                "horizontalAlignment": "CENTER",
                                "verticalAlignment": "MIDDLE",
                                "textFormat": {
                                    "bold": True,
                                    "foregroundColor": {"red":0,"green":0,"blue":0}
                                }
                            }
                        },
                        "fields": "userEnteredFormat(backgroundColor,horizontalAlignment,verticalAlignment,textFormat)"
                    }
                })

                # (D) 합계행 병합 + 서식
                sum_row_0based = row_cursor_fluxus_detail_end - 1
                fluxus_detail_requests.append({
                    "mergeCells": {
                        "range": {
                            "sheetId": sheet_id_fluxus_detail,
                            "startRowIndex": sum_row_0based,
                            "endRowIndex": sum_row_0based+1,
                            "startColumnIndex": 0,
                            "endColumnIndex": 6
                        },
                        "mergeType": "MERGE_ALL"
                    }
                })
                fluxus_detail_requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": sheet_id_fluxus_detail,
                            "startRowIndex": sum_row_0based,
                            "endRowIndex": sum_row_0based+1,
                            "startColumnIndex": 0,
                            "endColumnIndex": 6
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "backgroundColor": {"red": 0.3, "green": 0.82, "blue": 0.88},
                                "horizontalAlignment": "CENTER",
                                "verticalAlignment": "MIDDLE",
                                "textFormat": {"bold": True}
                            }
                        },
                        "fields": "userEnteredFormat(backgroundColor,horizontalAlignment,verticalAlignment,textFormat)"
                    }
                })
                # 합계값(G열)에 오른쪽 정렬
                fluxus_detail_requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": sheet_id_fluxus_detail,
                            "startRowIndex": sum_row_0based,
                            "endRowIndex": sum_row_0based+1,
                            "startColumnIndex": 6,
                            "endColumnIndex": 7
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "horizontalAlignment": "RIGHT",
                                "textFormat": {"bold": True}
                            }
                        },
                        "fields": "userEnteredFormat(horizontalAlignment,textFormat)"
                    }
                })
                # 매출 순수익 칼럼 (F열=idx=6) 나머지 행들
                fluxus_detail_requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": sheet_id_fluxus_detail,
                            "startRowIndex": 1,
                            "endRowIndex": sum_row_0based,
                            "startColumnIndex": 6,
                            "endColumnIndex": 7
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "horizontalAlignment": "RIGHT"
                            }
                        },
                        "fields": "userEnteredFormat(horizontalAlignment)"
                    }
                })

                # (E) 전체 테두리
                fluxus_detail_requests.append({
                    "updateBorders": {
                        "range": {
                            "sheetId": sheet_id_fluxus_detail,
                            "startRowIndex": 0,
                            "endRowIndex": row_cursor_fluxus_detail_end,
                            "startColumnIndex": 0,
                            "endColumnIndex": 7
                        },
                        "top":    {"style":"SOLID","width":1},
                        "bottom": {"style":"SOLID","width":1},
                        "left":   {"style":"SOLID","width":1},
                        "right":  {"style":"SOLID","width":1},
                        "innerHorizontal": {"style":"SOLID","width":1},
                        "innerVertical":   {"style":"SOLID","width":1}
                    }
                })

                all_requests.extend(fluxus_detail_requests)

                # (호출 횟수 분할) 1회 batchUpdate 요청이 너무 커지면 나눠서 전송
                if len(all_requests) >= 200:
                    sheet_svc.spreadsheets().batchUpdate(
                        spreadsheetId=out_file_id,
                        body={"requests": all_requests}
                    ).execute()
                    all_requests.clear()
                    time.sleep(1)
            

                # ################################
                # FLUXUS 정산서 탭 (batchUpdate 방식)
                # ################################
                ws_fluxus_report = out_sh.worksheet(f"{one_sosok}_{artist}(정산서)")
                ws_fluxus_report_id = ws_fluxus_report.id

                # 매출 합
                fluxus_sum_1 = sum(d["revenue"] for d in fluxus_yt_details_sorted)  # "음원서비스별" 총합
                fluxus_sum_2 = sum(d["revenue"] for d in fluxus_fs_details_sorted)
                fluxus_sum_all = fluxus_sum_1 + fluxus_sum_2

                # 앨범별 합
                fluxus_album_sum = defaultdict(float)
                for d in fluxus_yt_details_sorted:
                    fluxus_album_sum[d["album"]] += d["revenue"]
                fluxus_sum_album = sum(fluxus_album_sum.values())

                # (A) "곡비" = "전월 잔액 + 당월 발생액" (요청 사항)
                prev_val = artist_cost_dict[artist]["전월잔액"]
                curr_val = artist_cost_dict[artist]["당월발생"]
                # 보고서 '3. 공제 내역'의 '곡비' 칼럼 값 = prev_val + curr_val
                song_cost_for_report = prev_val + curr_val

                # (B) 공제 금액 & 잔액
                deduct_val = artist_cost_dict[artist]["당월차감액"]  # 이미 input_song cost에서 계산된 값
                remain_val = artist_cost_dict[artist]["당월잔액"]   # 동일
                # "공제 적용 후" 매출 = (음원 매출 합) - 공제금액 => sum_2 - deduct_val
                # (단, 요청 사항/업무로직에 따라 정확히 어떻게 적용할지는 케이스별로 맞춤)

                # (C) 정산율 / 최종 정산금액
                rate_val = artist_cost_dict[artist]["정산요율"]
                공제적용후 = sum_2 - deduct_val
                final_amount = 공제적용후 * (rate_val / 100.0)
                

                # --------------------------------------
                # 정산서 테이블(직접 row col 배열 채우기)
                # --------------------------------------
                report_fluxus_matrix = []
                for _ in range(300):
                    report_fluxus_matrix.append([""] * 7)

                # 1) 상단 공통정보
                report_fluxus_matrix[1][5] = report_date   # 보고서 발행일
                report_fluxus_matrix[3][1] = f"{year_val}년 {month_val}월 판매분"
                report_fluxus_matrix[5][1] = f"{artist}님 음원 정산 내역서"

                report_fluxus_matrix[7][0] = "•"
                report_fluxus_matrix[7][1] = "저희와 함께해 주셔서 정말 감사하고 앞으로도 잘 부탁드리겠습니다!"
                report_fluxus_matrix[8][0] = "•"
                report_fluxus_matrix[8][1] = f"{year_val}년 {month_val}월 음원의 수익을 아래와 같이 정산드리오니 참고 부탁드립니다."
                report_fluxus_matrix[9][0] = "•"
                report_fluxus_matrix[9][1] = "정산 관련하여 문의사항이 있다면 무엇이든, 언제든 편히 메일 주세요!"
                report_fluxus_matrix[9][4] = "E-Mail : lucasdh3013@naver.com"

                # -----------------------------------------------------------------
                # 1. 음원 서비스별 정산내역 (세부매출 그대로)
                # -----------------------------------------------------------------
                report_fluxus_matrix[12][0] = "1."
                report_fluxus_matrix[12][1] = "음원 서비스별 정산내역"

                header_row_1 = 13
                headers_1 = ["앨범", "내용", "", "기간", "매출액"]
                for i_h, val_h in enumerate(headers_1):
                    report_fluxus_matrix[header_row_1][1 + i_h] = val_h

                row_cursor = header_row_1 + 1
                for d in fluxus_yt_details_sorted:
                    rv = d["revenue"]
                    report_fluxus_matrix[row_cursor][1] = d["album"]
                    report_fluxus_matrix[row_cursor][2] = d["track_title"]
                    report_fluxus_matrix[row_cursor][4] = f"{year_val}년 {month_val}월"
                    report_fluxus_matrix[row_cursor][5] = to_currency(rv)
                    row_cursor += 1
                    if len(fluxus_yt_details_sorted) == row_cursor-13:
                        report_fluxus_matrix[row_cursor][1] = d["album"]
                        report_fluxus_matrix[row_cursor][2] = f"국내, 해외 플랫폼({int(month_val)-1}월)"
                        report_fluxus_matrix[row_cursor][4] = f"{year_val}년 {month_val}월"
                        report_fluxus_matrix[row_cursor][5] = to_currency(sum_fs_rv_val)
                        fluxus_album_sum[d["album"]] += sum_fs_rv_val
                        row_cursor += 1

                row_cursor += 2
                # 합계
                report_fluxus_matrix[row_cursor-1][1] = "합계"
                report_fluxus_matrix[row_cursor-1][5] = to_currency(fluxus_sum_all)
                row_cursor_sum1 = row_cursor
                row_cursor += 1

                # -----------------------------------------------------------------
                # 2. 앨범 별 정산 내역
                # -----------------------------------------------------------------
                report_fluxus_matrix[row_cursor][0] = "2."
                report_fluxus_matrix[row_cursor][1] = "앨범 별 정산 내역"
                row_cursor += 1
                row_cursor_album = row_cursor
                report_fluxus_matrix[row_cursor][1] = "앨범"
                report_fluxus_matrix[row_cursor][4] = "기간"
                report_fluxus_matrix[row_cursor][5] = "매출액"
                row_cursor += 1

                for alb in sorted(fluxus_album_sum.keys(), key=album_sort_key):
                    amt = fluxus_album_sum[alb]
                    report_fluxus_matrix[row_cursor][1] = alb
                    report_fluxus_matrix[row_cursor][4] = f"{year_val}년 {month_val}월"
                    report_fluxus_matrix[row_cursor][5] = to_currency(amt)
                    row_cursor += 1

                row_cursor += 1
                report_fluxus_matrix[row_cursor-1][1] = "합계"
                report_fluxus_matrix[row_cursor-1][5] = to_currency(fluxus_sum_album)
                row_cursor_sum2 = row_cursor
                row_cursor += 1

                # -----------------------------------------------------------------
                # 3. 공제 내역
                #    (요청사항: '곡비' 칼럼 = (전월 잔액 + 당월 발생액))
                # -----------------------------------------------------------------
                report_fluxus_matrix[row_cursor][0] = "3."
                report_fluxus_matrix[row_cursor][1] = "공제 내역"
                row_cursor += 1
                row_cursor_deduction = row_cursor

                report_fluxus_matrix[row_cursor][1] = "앨범"
                report_fluxus_matrix[row_cursor][2] = "곡비"
                report_fluxus_matrix[row_cursor][3] = "공제 금액"
                report_fluxus_matrix[row_cursor][4] = "공제 후 남은 곡비"
                report_fluxus_matrix[row_cursor][5] = "공제 적용 금액"
                row_cursor += 1

                # 앨범(들)을 표기만 할지, 혹은 여러줄로 표현할지 등은 업무 규칙에 따라
                alb_list = sorted(fluxus_album_sum.keys(), key=album_sort_key)
                alb_str = ", ".join(alb_list) if alb_list else "(앨범 없음)"

                report_fluxus_matrix[row_cursor][1] = alb_str
                # (중요) 여기서 "곡비" = prev_val + curr_val
                report_fluxus_matrix[row_cursor][2] = to_currency(song_cost_for_report)
                # 공제금액
                report_fluxus_matrix[row_cursor][3] = to_currency(deduct_val)
                # 공제 후 남은 곡비
                report_fluxus_matrix[row_cursor][4] = to_currency(remain_val)
                # 공제 적용 금액 (매출 - 공제금액)
                report_fluxus_matrix[row_cursor][5] = to_currency(fluxus_sum_all - deduct_val)
                row_cursor += 2
                row_cursor_sum3 = row_cursor

                # -----------------------------------------------------------------
                # 4. 수익 배분
                # -----------------------------------------------------------------
                report_fluxus_matrix[row_cursor][0] = "4."
                report_fluxus_matrix[row_cursor][1] = "수익 배분"
                row_cursor += 1
                row_cursor_rate = row_cursor
                report_fluxus_matrix[row_cursor][1] = "앨범"
                report_fluxus_matrix[row_cursor][2] = "항목"
                report_fluxus_matrix[row_cursor][3] = "적용율"
                report_fluxus_matrix[row_cursor][5] = "적용 금액"
                row_cursor += 1

                report_fluxus_matrix[row_cursor][1] = alb_str
                report_fluxus_matrix[row_cursor][2] = "수익 배분율"
                report_fluxus_matrix[row_cursor][3] = f"{int(rate_val)}%"
                report_fluxus_matrix[row_cursor][5] = to_currency(final_amount)
                row_cursor += 1

                report_fluxus_matrix[row_cursor][1] = "총 정산금액"
                report_fluxus_matrix[row_cursor][5] = to_currency(final_amount)
                row_cursor_sum4 = row_cursor
                row_cursor += 2

                report_fluxus_matrix[row_cursor][5] = "* 부가세 별도"
                row_cursor_report_end = row_cursor + 2

                # 시트에 실제 업로드
                ws_fluxus_report.update("A1", report_fluxus_matrix)
                time.sleep(1)

                # ------------------------------------
                # (검증) check_dict에 비교결과 반영
                # ------------------------------------
                # (1) 세부매출 vs 정산서
                for d in fluxus_yt_details_sorted:
                    original_val = d["revenue"]
                    report_val   = d["revenue"]  # 현재는 동일
                    is_match = almost_equal(original_val, report_val)
                    if not is_match:
                        check_dict["verification_summary"]["total_errors"] += 1
                        check_dict["verification_summary"]["artist_error_list"].append(artist)

                    row_report_item = {
                        "아티스트": artist,
                        "앨범": d["album"],
                        "타이틀명": d["track_title"],
                        "원본_매출액": original_val,
                        "정산서_매출액": report_val,
                        "match_매출액": is_match,
                    }
                    check_dict["details_verification"]["세부매출"].append(row_report_item)

                # (2) 공제 내역(곡비,공제금액,공제후잔액)
                #   원본(= input_song cost) 값 vs 보고서 값
                #   "곡비"는 (prev + curr), "공제금액"=deduct_val, "남은 곡비"=remain_val
                #   *원본_곡비 = (전월잔액 + 당월발생)
                original_song_cost = artist_cost_dict[artist]["전월잔액"] + artist_cost_dict[artist]["당월발생"]
                is_match_songcost = almost_equal(original_song_cost, song_cost_for_report)
                is_match_deduct   = almost_equal(artist_cost_dict[artist]["당월차감액"], deduct_val)
                is_match_remain   = almost_equal(artist_cost_dict[artist]["당월잔액"], remain_val)

                if not (is_match_songcost and is_match_deduct and is_match_remain):
                    check_dict["verification_summary"]["total_errors"] += 1
                    check_dict["verification_summary"]["artist_error_list"].append(artist)

                row_report_item_3 = {
                    "아티스트": artist,
                    "구분": "공제내역",
                    # 곡비
                    "원본_곡비": original_song_cost,
                    "정산서_곡비": song_cost_for_report,
                    "match_곡비": is_match_songcost,
                    # 공제금액
                    "원본_공제금액": artist_cost_dict[artist]["당월차감액"],
                    "정산서_공제금액": deduct_val,
                    "match_공제금액": is_match_deduct,
                    # 공제후잔액
                    "원본_공제후잔액": artist_cost_dict[artist]["당월잔액"],
                    "정산서_공제후잔액": remain_val,
                    "match_공제후잔액": is_match_remain,
                }
                check_dict["details_verification"]["정산서"].append(row_report_item_3)

                # (3) 4번 수익 배분율
                original_rate = artist_cost_dict[artist]["정산요율"]
                is_rate_match = almost_equal(original_rate, rate_val)
                if not is_rate_match:
                    check_dict["verification_summary"]["total_errors"] += 1
                    check_dict["verification_summary"]["artist_error_list"].append(artist)

                row_report_item_4 = {
                    "아티스트": artist,
                    "구분": "수익배분율",
                    "원본_정산율(%)": original_rate,
                    "정산서_정산율(%)": rate_val,
                    "match_정산율": is_rate_match,
                }
                check_dict["details_verification"]["정산서"].append(row_report_item_4)

                time.sleep(1)   

                # --------------------------------------------------
                # 정산서 탭(디자인/서식) batchUpdate
                # --------------------------------------------------
                report_fluxus_requests = []

                # (A) 시트 row/col 크기
                report_fluxus_requests.append({
                    "updateSheetProperties": {
                        "properties": {
                            "sheetId": ws_fluxus_report_id,
                            "gridProperties": {
                                "rowCount": row_cursor_report_end,
                                "columnCount": 7
                            }
                        },
                        "fields": "gridProperties(rowCount,columnCount)"
                    }
                })

                # (B) 열너비 (A=0 ~ H=7)
                report_fluxus_requests.extend([
                    {
                        "updateDimensionProperties": {
                            "range": {
                                "sheetId": ws_fluxus_report_id,
                                "dimension": "COLUMNS",
                                "startIndex": 0,
                                "endIndex": 1
                            },
                            "properties": {"pixelSize": 40},
                            "fields": "pixelSize"
                        }
                    },
                    {
                        "updateDimensionProperties": {
                            "range": {
                                "sheetId": ws_fluxus_report_id,
                                "dimension": "COLUMNS",
                                "startIndex": 1,
                                "endIndex": 2
                            },
                            "properties": {"pixelSize": 180},
                            "fields": "pixelSize"
                        }
                    },
                    {
                        "updateDimensionProperties": {
                            "range": {
                                "sheetId": ws_fluxus_report_id,
                                "dimension": "COLUMNS",
                                "startIndex": 2,
                                "endIndex": 3
                            },
                            "properties": {"pixelSize": 160},
                            "fields": "pixelSize"
                        }
                    },
                    {
                        "updateDimensionProperties": {
                            "range": {
                                "sheetId": ws_fluxus_report_id,
                                "dimension": "COLUMNS",
                                "startIndex": 3,
                                "endIndex": 4
                            },
                            "properties": {"pixelSize": 140},
                            "fields": "pixelSize"
                        }
                    },
                    {
                        "updateDimensionProperties": {
                            "range": {
                                "sheetId": ws_fluxus_report_id,
                                "dimension": "COLUMNS",
                                "startIndex": 4,
                                "endIndex": 5
                            },
                            "properties": {"pixelSize": 150},
                            "fields": "pixelSize"
                        }
                    },
                    {
                        "updateDimensionProperties": {
                            "range": {
                                "sheetId": ws_fluxus_report_id,
                                "dimension": "COLUMNS",
                                "startIndex": 5,
                                "endIndex": 6
                            },
                            "properties": {"pixelSize": 160},
                            "fields": "pixelSize"
                        }
                    },
                    {
                        "updateDimensionProperties": {
                            "range": {
                                "sheetId": ws_fluxus_report_id,
                                "dimension": "COLUMNS",
                                "startIndex": 6,
                                "endIndex": 7
                            },
                            "properties": {"pixelSize": 40},
                            "fields": "pixelSize"
                        }
                    },
                ])

                # (C) 특정행 높이
                report_fluxus_requests.extend([
                    {
                        "updateDimensionProperties": {
                            "range": {
                                "sheetId": ws_fluxus_report_id,
                                "dimension": "ROWS",
                                "startIndex": 3,
                                "endIndex": 4
                            },
                            "properties": {"pixelSize": 30},
                            "fields": "pixelSize"
                        }
                    },
                    {
                        "updateDimensionProperties": {
                            "range": {
                                "sheetId": ws_fluxus_report_id,
                                "dimension": "ROWS",
                                "startIndex": 5,
                                "endIndex": 6
                            },
                            "properties": {"pixelSize": 30},
                            "fields": "pixelSize"
                        }
                    },
                ])

                # (D) 상단 고정 항목(발행 날짜, H2: row=1, col=6)
                report_fluxus_requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": ws_fluxus_report_id,
                            "startRowIndex": 1,
                            "endRowIndex": 2,
                            "startColumnIndex": 5,
                            "endColumnIndex": 6
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "horizontalAlignment": "RIGHT",
                                "verticalAlignment": "MIDDLE",
                                "textFormat": {
                                    "fontFamily": "Malgun Gothic",
                                    "fontSize": 10,
                                    "bold": False
                                }
                            }
                        },
                        "fields": "userEnteredFormat(horizontalAlignment,verticalAlignment,textFormat)"
                    }
                })        

                # (E) 상단 고정 항목(판매분, B4:E4)
                report_fluxus_requests.append({
                    "mergeCells": {
                        "range": {
                            "sheetId": ws_fluxus_report_id,
                            "startRowIndex": 3,  # (4-1)
                            "endRowIndex": 4,
                            "startColumnIndex": 1,  # (B=1)
                            "endColumnIndex": 4     # (E=4 => endIndex=5)
                        },
                        "mergeType": "MERGE_ALL"
                    }
                })
                report_fluxus_requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": ws_fluxus_report_id,
                            "startRowIndex": 3,
                            "endRowIndex": 4,
                            "startColumnIndex": 1,
                            "endColumnIndex": 4
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "horizontalAlignment": "LEFT",
                                "verticalAlignment": "MIDDLE",
                                "textFormat": {
                                    "fontFamily": "Malgun Gothic",
                                    "fontSize": 15,
                                    "bold": True
                                }
                            }
                        },
                        "fields": "userEnteredFormat(horizontalAlignment,verticalAlignment,textFormat)"
                    }
                })

                # (F) 상단 고정 항목(아티스트 정산내역서, B6:G6)
                report_fluxus_requests.append({
                    "mergeCells": {
                        "range": {
                            "sheetId": ws_fluxus_report_id,
                            "startRowIndex": 5,
                            "endRowIndex": 6,
                            "startColumnIndex": 1,
                            "endColumnIndex": 6
                        },
                        "mergeType": "MERGE_ALL"
                    }
                })
                report_fluxus_requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": ws_fluxus_report_id,
                            "startRowIndex": 5,
                            "endRowIndex": 6,
                            "startColumnIndex": 1,
                            "endColumnIndex": 6
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "backgroundColor": {"red": 0.896, "green": 0.988, "blue": 1},
                                "horizontalAlignment": "CENTER",
                                "verticalAlignment": "MIDDLE",
                                "textFormat": {
                                    "fontFamily": "Malgun Gothic",
                                    "fontSize": 15,
                                    "bold": True
                                }
                            }
                        },
                        "fields": "userEnteredFormat(backgroundColor,horizontalAlignment,verticalAlignment,textFormat)"
                    }
                })

                # (G) 상단 고정 항목(안내문, B8:E8~B10:E10)
                #8행
                report_fluxus_requests.append({
                    "mergeCells": {
                        "range": {
                            "sheetId": ws_fluxus_report_id,
                            "startRowIndex": 7,  # (4-1)
                            "endRowIndex": 8,
                            "startColumnIndex": 1,  # (B=1)
                            "endColumnIndex": 4     # (E=4 => endIndex=5)
                        },
                        "mergeType": "MERGE_ALL"
                    }
                })
                report_fluxus_requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": ws_fluxus_report_id,
                            "startRowIndex": 7,  # (4-1)
                            "endRowIndex": 8,
                            "startColumnIndex": 1,  # (B=1)
                            "endColumnIndex": 4     # (E=4 => endIndex=5)
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "horizontalAlignment": "LEFT",
                                "verticalAlignment": "MIDDLE",
                                "textFormat": {
                                    "bold": False
                                }
                            }
                        },
                        "fields": "userEnteredFormat(horizontalAlignment,verticalAlignment,textFormat)"
                    }
                })
                #9행
                report_fluxus_requests.append({
                    "mergeCells": {
                        "range": {
                            "sheetId": ws_fluxus_report_id,
                            "startRowIndex": 8,  
                            "endRowIndex": 9,
                            "startColumnIndex": 1,
                            "endColumnIndex": 4 
                        },
                        "mergeType": "MERGE_ALL"
                    }
                })
                report_fluxus_requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": ws_fluxus_report_id,
                            "startRowIndex": 8,  
                            "endRowIndex": 9,
                            "startColumnIndex": 1,
                            "endColumnIndex": 4 
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "horizontalAlignment": "LEFT",
                                "verticalAlignment": "MIDDLE",
                                "textFormat": {
                                "bold": False
                                }
                            }
                        },
                        "fields": "userEnteredFormat(horizontalAlignment,verticalAlignment,textFormat)"
                    }
                })
                #10행
                report_fluxus_requests.append({
                    "mergeCells": {
                        "range": {
                            "sheetId": ws_fluxus_report_id,
                            "startRowIndex": 9,  
                            "endRowIndex": 10,
                            "startColumnIndex": 1,
                            "endColumnIndex": 4 
                        },
                        "mergeType": "MERGE_ALL"
                    }
                })
                report_fluxus_requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": ws_fluxus_report_id,
                            "startRowIndex": 9,  
                            "endRowIndex": 10,
                            "startColumnIndex": 1,
                            "endColumnIndex": 4 
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "horizontalAlignment": "LEFT",
                                "verticalAlignment": "MIDDLE",
                                "textFormat": {
                                    "bold": False
                                }
                            }
                        },
                        "fields": "userEnteredFormat(horizontalAlignment,verticalAlignment,textFormat)"
                    }
                })
                # 10행 (E-Mail 칸)
                report_fluxus_requests.append({
                    "mergeCells": {
                        "range": {
                            "sheetId": ws_fluxus_report_id,
                            "startRowIndex": 9,  
                            "endRowIndex": 10,
                            "startColumnIndex": 4,
                            "endColumnIndex": 6 
                        },
                        "mergeType": "MERGE_ALL"
                    }
                })
                report_fluxus_requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": ws_fluxus_report_id,
                            "startRowIndex": 9,
                            "endRowIndex": 10,
                            "startColumnIndex": 4,
                            "endColumnIndex": 6
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "horizontalAlignment": "RIGHT",
                                "verticalAlignment": "MIDDLE",
                                "textFormat": {
                                    "foregroundColor": {"red": 0.29, "green": 0.53, "blue": 0.91},
                                    "fontFamily": "Malgun Gothic",
                                    "fontSize": 10,
                                    "bold": True
                                }
                            }
                        },
                        "fields": "userEnteredFormat(horizontalAlignment,verticalAlignment,textFormat)"
                    }
                })
                
                # (H) 1열 정렬 (번호 영역)
                report_fluxus_requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": ws_fluxus_report_id,
                            "startRowIndex": 1,
                            "endRowIndex": row_cursor_rate+1,
                            "startColumnIndex": 0,
                            "endColumnIndex": 1
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "horizontalAlignment": "CENTER",
                                "verticalAlignment": "MIDDLE",
                                "textFormat": {
                                    "fontFamily": "Malgun Gothic",
                                    "fontSize": 10,
                                    "bold": False
                                }
                            }
                        },
                        "fields": "userEnteredFormat(horizontalAlignment,verticalAlignment,textFormat)"
                    }
                })

                # (I) 하단 고정 항목(부가세, G)
                report_fluxus_requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": ws_fluxus_report_id,
                            "startRowIndex": row_cursor_report_end-2,
                            "endRowIndex": row_cursor_report_end,
                            "startColumnIndex": 5,
                            "endColumnIndex": 6
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "horizontalAlignment": "RIGHT",
                                "verticalAlignment": "MIDDLE",
                                "textFormat": {
                                    "fontFamily": "Malgun Gothic",
                                    "fontSize": 10,
                                    "bold": False
                                }
                            }
                        },
                        "fields": "userEnteredFormat(horizontalAlignment,verticalAlignment,textFormat)"
                    }
                }) 
            

                # (J-1) "음원 서비스별 정산내역" 표 타이틀
                report_fluxus_requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": ws_fluxus_report_id,
                            "startRowIndex": 12,  # (4-1)
                            "endRowIndex": 13,
                            "startColumnIndex": 1,  # (B=1)
                            "endColumnIndex": 2     # (E=4 => endIndex=5)
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "horizontalAlignment": "LEFT",
                                "verticalAlignment": "MIDDLE",
                                "textFormat": {
                                    "bold": True
                                }
                            }
                        },
                        "fields": "userEnteredFormat(horizontalAlignment,verticalAlignment,textFormat)"
                    }
                })
                # (J-2) "음원 서비스별 정산내역" 표 헤더 (Row=13)
                report_fluxus_requests.append({
                    "mergeCells": {
                        "range": {
                            "sheetId": ws_fluxus_report_id,
                            "startRowIndex": 13,
                            "endRowIndex": 14,
                            "startColumnIndex": 2,
                            "endColumnIndex": 4
                        },
                        "mergeType": "MERGE_ALL"
                    }
                })
                report_fluxus_requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": ws_fluxus_report_id,
                            "startRowIndex": 13,
                            "endRowIndex": 14,
                            "startColumnIndex": 1,
                            "endColumnIndex": 6
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "backgroundColor": {"red": 0.3, "green": 0.82, "blue": 0.88},
                                "horizontalAlignment": "CENTER",
                                "verticalAlignment": "MIDDLE",
                                "textFormat": {
                                    "fontFamily": "Malgun Gothic",
                                    "fontSize": 10,
                                    "bold": True
                                }
                            }
                        },
                        "fields": "userEnteredFormat(backgroundColor,horizontalAlignment,verticalAlignment,textFormat)"
                    }
                })
                # (J-3) 합계행 전 병합
                report_fluxus_requests.append({
                    "mergeCells": {
                        "range": {
                            "sheetId": ws_fluxus_report_id,
                            "startRowIndex": row_cursor_sum1-2,
                            "endRowIndex": row_cursor_sum1-1,
                            "startColumnIndex": 1,
                            "endColumnIndex": 6
                        },
                        "mergeType": "MERGE_ALL"
                    }
                })
                # (J-4) 합계행 병합
                report_fluxus_requests.append({
                    "mergeCells": {
                        "range": {
                            "sheetId": ws_fluxus_report_id,
                            "startRowIndex": row_cursor_sum1-1,
                            "endRowIndex": row_cursor_sum1,
                            "startColumnIndex": 1,
                            "endColumnIndex": 5
                        },
                        "mergeType": "MERGE_ALL"
                    }
                })
                report_fluxus_requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": ws_fluxus_report_id,
                            "startRowIndex": row_cursor_sum1-1,
                            "endRowIndex": row_cursor_sum1,
                            "startColumnIndex": 1,
                            "endColumnIndex": 5
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "backgroundColor": {"red": 0.896, "green": 0.988, "blue": 1},
                                "horizontalAlignment": "CENTER",
                                "verticalAlignment": "MIDDLE",
                                "textFormat": {
                                    "fontFamily": "Malgun Gothic",
                                    "fontSize": 10,
                                    "bold": True
                                }
                            }
                        },
                        "fields": "userEnteredFormat(backgroundColor,horizontalAlignment,verticalAlignment,textFormat)"
                    }
                })
                report_fluxus_requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": ws_fluxus_report_id,
                            "startRowIndex": row_cursor_sum1-1,
                            "endRowIndex": row_cursor_sum1,
                            "startColumnIndex": 5,
                            "endColumnIndex": 6
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "backgroundColor": {"red": 0.896, "green": 0.988, "blue": 1},
                                "horizontalAlignment": "CENTER",
                                "verticalAlignment": "MIDDLE",
                                "textFormat": {
                                    "fontFamily": "Malgun Gothic",
                                    "fontSize": 10,
                                    "bold": True
                                }
                            }
                        },
                        "fields": "userEnteredFormat(backgroundColor,horizontalAlignment,verticalAlignment,textFormat)"
                    }
                })
                # (J-5) 표에 Banding (줄무늬 효과)
                banding_start_row = 14
                banding_end_row = row_cursor_sum1 - 2
                banding_start_col = 1
                banding_end_col = 6
                if banding_end_row > banding_start_row:  # 유효범위 체크
                    report_fluxus_requests.append({
                        "addBanding": {
                            "bandedRange": {
                                "range": {
                                    "sheetId": ws_fluxus_report_id,
                                    "startRowIndex": banding_start_row,
                                    "endRowIndex": banding_end_row,
                                    "startColumnIndex": banding_start_col,
                                    "endColumnIndex": banding_end_col
                                },
                                "rowProperties": {
                                    "firstBandColor": {
                                        "red": 1.0, "green": 1.0, "blue": 1.0
                                    },
                                    "secondBandColor": {
                                        "red": 0.896, "green": 0.988, "blue": 1
                                    }
                                },
                                
                            }
                        }
                    })
                    report_fluxus_requests.append({
                        "repeatCell": {
                            "range": {
                                "sheetId": ws_fluxus_report_id,
                                "startRowIndex": banding_start_row,
                                "endRowIndex": banding_end_row,
                                "startColumnIndex": banding_start_col,
                                "endColumnIndex": banding_end_col
                            },
                            "cell": {
                                "userEnteredFormat": {
                                    "horizontalAlignment": "CENTER",
                                    "verticalAlignment": "MIDDLE",
                                    "textFormat": {
                                        "fontFamily": "Malgun Gothic",
                                        "fontSize": 10,
                                        "bold": False
                                    }
                                }
                            },
                            "fields": "userEnteredFormat(horizontalAlignment,verticalAlignment,textFormat)"
                        }
                    })


                # (K-1) 앨범별 정산내역 타이틀
                report_fluxus_requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": ws_fluxus_report_id,
                            "startRowIndex": row_cursor_album-1,
                            "endRowIndex": row_cursor_album,
                            "startColumnIndex": 1, 
                            "endColumnIndex": 2    
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "horizontalAlignment": "LEFT",
                                "verticalAlignment": "MIDDLE",
                                "textFormat": {
                                    "bold": True
                                }
                            }
                        },
                        "fields": "userEnteredFormat(horizontalAlignment,verticalAlignment,textFormat)"
                    }
                })
                # (K-2) 앨범별 정산내역 헤더
                report_fluxus_requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": ws_fluxus_report_id,
                            "startRowIndex": row_cursor_album,
                            "endRowIndex": row_cursor_album+1,
                            "startColumnIndex": 1,
                            "endColumnIndex": 6
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "backgroundColor": {"red": 0.3, "green": 0.82, "blue": 0.88},
                                "horizontalAlignment": "CENTER",
                                "verticalAlignment": "MIDDLE",
                                "textFormat": {
                                    "fontFamily": "Malgun Gothic",
                                    "fontSize": 10,
                                    "bold": True
                                }
                            }
                        },
                        "fields": "userEnteredFormat(backgroundColor,horizontalAlignment,verticalAlignment,textFormat)"
                    }
                })
                # (K-3) 앨범별 정산내역 표 본문
                report_fluxus_requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": ws_fluxus_report_id,
                            "startRowIndex": row_cursor_album+1,
                            "endRowIndex": row_cursor_sum2-1,
                            "startColumnIndex": 1,
                            "endColumnIndex": 6
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "horizontalAlignment": "CENTER",
                                "verticalAlignment": "MIDDLE",
                                "textFormat": {
                                    "fontFamily": "Malgun Gothic",
                                    "fontSize": 10,
                                    "bold": False
                                }
                            }
                        },
                        "fields": "userEnteredFormat(horizontalAlignment,verticalAlignment,textFormat)"
                    }
                })
                # (K-4) 앨범별 정산내역 합계행
                report_fluxus_requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": ws_fluxus_report_id,
                            "startRowIndex": row_cursor_sum2-1,
                            "endRowIndex": row_cursor_sum2,
                            "startColumnIndex": 1,
                            "endColumnIndex": 6
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "horizontalAlignment": "CENTER",
                                "verticalAlignment": "MIDDLE",
                                "textFormat": {
                                    "fontFamily": "Malgun Gothic",
                                    "fontSize": 10,
                                    "bold": True
                                }
                            }
                        },
                        "fields": "userEnteredFormat(horizontalAlignment,verticalAlignment,textFormat)"
                    }
                })
                # (K-5) 합계행 병합
                report_fluxus_requests.append({
                    "mergeCells": {
                        "range": {
                            "sheetId": ws_fluxus_report_id,
                            "startRowIndex": row_cursor_sum2-1,
                            "endRowIndex": row_cursor_sum2,
                            "startColumnIndex": 1,
                            "endColumnIndex": 5
                        },
                        "mergeType": "MERGE_ALL"
                    }
                })
                report_fluxus_requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": ws_fluxus_report_id,
                            "startRowIndex": row_cursor_sum2-1,
                            "endRowIndex": row_cursor_sum2,
                            "startColumnIndex": 1,
                            "endColumnIndex": 5
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "backgroundColor": {"red": 0.896, "green": 0.988, "blue": 1},
                                "horizontalAlignment": "CENTER",
                                "verticalAlignment": "MIDDLE",
                                "textFormat": {
                                    "fontFamily": "Malgun Gothic",
                                    "fontSize": 10,
                                    "bold": True
                                }
                            }
                        },
                        "fields": "userEnteredFormat(backgroundColor,horizontalAlignment,verticalAlignment,textFormat)"
                    }
                })
                report_fluxus_requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": ws_fluxus_report_id,
                            "startRowIndex": row_cursor_sum2-1,
                            "endRowIndex": row_cursor_sum2,
                            "startColumnIndex": 5,
                            "endColumnIndex": 6
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "backgroundColor": {"red": 0.896, "green": 0.988, "blue": 1},
                                "horizontalAlignment": "CENTER",
                                "verticalAlignment": "MIDDLE",
                                "textFormat": {
                                    "fontFamily": "Malgun Gothic",
                                    "fontSize": 10,
                                    "bold": True
                                }
                            }
                        },
                        "fields": "userEnteredFormat(backgroundColor,horizontalAlignment,verticalAlignment,textFormat)"
                    }
                })


                # (L-1) 공제 내역 타이틀
                report_fluxus_requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": ws_fluxus_report_id,
                            "startRowIndex": row_cursor_deduction-1,  # (4-1)
                            "endRowIndex": row_cursor_deduction,
                            "startColumnIndex": 1,  # (B=1)
                            "endColumnIndex": 2     # (E=4 => endIndex=5)
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "horizontalAlignment": "LEFT",
                                "verticalAlignment": "MIDDLE",
                                "textFormat": {
                                    "bold": True
                                }
                            }
                        },
                        "fields": "userEnteredFormat(horizontalAlignment,verticalAlignment,textFormat)"
                    }
                })
                # (L-2) 공제 내역 헤더
                report_fluxus_requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": ws_fluxus_report_id,
                            "startRowIndex": row_cursor_deduction,
                            "endRowIndex": row_cursor_deduction+1,
                            "startColumnIndex": 1,
                            "endColumnIndex": 6
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "backgroundColor": {"red": 0.3, "green": 0.82, "blue": 0.88},
                                "horizontalAlignment": "CENTER",
                                "verticalAlignment": "MIDDLE",
                                "textFormat": {
                                    "fontFamily": "Malgun Gothic",
                                    "fontSize": 10,
                                    "bold": True
                                }
                            }
                        },
                        "fields": "userEnteredFormat(backgroundColor,horizontalAlignment,verticalAlignment,textFormat)"
                    }
                })
                # (L-3) 공제 내역 표 본문 (데이터부분)
                report_fluxus_requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": ws_fluxus_report_id,
                            "startRowIndex": row_cursor_deduction+1,
                            "endRowIndex": row_cursor_deduction+2,
                            "startColumnIndex": 1,
                            "endColumnIndex": 5
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "horizontalAlignment": "CENTER",
                                "verticalAlignment": "MIDDLE",
                                "textFormat": {
                                    "fontFamily": "Malgun Gothic",
                                    "fontSize": 10,
                                    "bold": False
                                }
                            }
                        },
                        "fields": "userEnteredFormat(horizontalAlignment,verticalAlignment,textFormat)"
                    }
                })
                # (L-4) 공제 내역 표 본문 (합계 부분)
                report_fluxus_requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": ws_fluxus_report_id,
                            "startRowIndex": row_cursor_deduction+1,
                            "endRowIndex": row_cursor_deduction+2,
                            "startColumnIndex": 5,
                            "endColumnIndex": 6
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "horizontalAlignment": "CENTER",
                                "verticalAlignment": "MIDDLE",
                                "textFormat": {
                                    "fontFamily": "Malgun Gothic",
                                    "fontSize": 10,
                                    "bold": True
                                }
                            }
                        },
                        "fields": "userEnteredFormat(horizontalAlignment,verticalAlignment,textFormat)"
                    }
                })


                # (M-1) 수익 배분 타이틀
                report_fluxus_requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": ws_fluxus_report_id,
                            "startRowIndex": row_cursor_rate-1,
                            "endRowIndex": row_cursor_rate,
                            "startColumnIndex": 1,  
                            "endColumnIndex": 2    
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "horizontalAlignment": "LEFT",
                                "verticalAlignment": "MIDDLE",
                                "textFormat": {
                                    "bold": True
                                }
                            }
                        },
                        "fields": "userEnteredFormat(horizontalAlignment,verticalAlignment,textFormat)"
                    }
                })
                # (M-2) 수익 배분 헤더
                report_fluxus_requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": ws_fluxus_report_id,
                            "startRowIndex": row_cursor_rate,
                            "endRowIndex": row_cursor_rate+1,
                            "startColumnIndex": 1,
                            "endColumnIndex": 6
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "backgroundColor": {"red": 0.3, "green": 0.82, "blue": 0.88},
                                "horizontalAlignment": "CENTER",
                                "verticalAlignment": "MIDDLE",
                                "textFormat": {
                                    "fontFamily": "Malgun Gothic",
                                    "fontSize": 10,
                                    "bold": True
                                }
                            }
                        },
                        "fields": "userEnteredFormat(backgroundColor,horizontalAlignment,verticalAlignment,textFormat)"
                    }
                })
                # (M-3) 수익 배분 표 본문 
                report_fluxus_requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": ws_fluxus_report_id,
                            "startRowIndex": row_cursor_rate+1,
                            "endRowIndex": row_cursor_rate+2,
                            "startColumnIndex": 1,
                            "endColumnIndex": 6
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "horizontalAlignment": "CENTER",
                                "verticalAlignment": "MIDDLE",
                                "textFormat": {
                                    "fontFamily": "Malgun Gothic",
                                    "fontSize": 10,
                                    "bold": False
                                }
                            }
                        },
                        "fields": "userEnteredFormat(horizontalAlignment,verticalAlignment,textFormat)"
                    }
                })
                # (M-4) 수익 배분 표 합계행 병합
                report_fluxus_requests.append({
                    "mergeCells": {
                        "range": {
                            "sheetId": ws_fluxus_report_id,
                            "startRowIndex": row_cursor_sum4,
                            "endRowIndex": row_cursor_sum4+1,
                            "startColumnIndex": 1,
                            "endColumnIndex": 5
                        },
                        "mergeType": "MERGE_ALL"
                    }
                })
                report_fluxus_requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": ws_fluxus_report_id,
                            "startRowIndex": row_cursor_sum4,
                            "endRowIndex": row_cursor_sum4+1,
                            "startColumnIndex": 1,
                            "endColumnIndex": 5
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "backgroundColor": {"red": 0.896, "green": 0.988, "blue": 1},
                                "horizontalAlignment": "CENTER",
                                "verticalAlignment": "MIDDLE",
                                "textFormat": {
                                    "fontFamily": "Malgun Gothic",
                                    "fontSize": 10,
                                    "bold": True
                                }
                            }
                        },
                        "fields": "userEnteredFormat(backgroundColor,horizontalAlignment,verticalAlignment,textFormat)"
                    }
                })
                report_fluxus_requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": ws_fluxus_report_id,
                            "startRowIndex": row_cursor_sum4,
                            "endRowIndex": row_cursor_sum4+1,
                            "startColumnIndex": 5,
                            "endColumnIndex": 6
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "backgroundColor": {"red": 0.896, "green": 0.988, "blue": 1},
                                "horizontalAlignment": "CENTER",
                                "verticalAlignment": "MIDDLE",
                                "textFormat": {
                                    "fontFamily": "Malgun Gothic",
                                    "fontSize": 10,
                                    "bold": True
                                }
                            }
                        },
                        "fields": "userEnteredFormat(backgroundColor,horizontalAlignment,verticalAlignment,textFormat)"
                    }
                })


                # (N) 전체 테두리 화이트
                report_fluxus_requests.append({
                    "updateBorders": {
                        "range": {
                            "sheetId": ws_fluxus_report_id,
                            "startRowIndex": 0,
                            "endRowIndex": row_cursor_report_end,
                            "startColumnIndex": 0,
                            "endColumnIndex": 7
                        },
                        "top":    {"style": "SOLID","width":1, "color":{"red":1,"green":1,"blue":1}},
                        "bottom": {"style": "SOLID","width":1, "color":{"red":1,"green":1,"blue":1}},
                        "left":   {"style": "SOLID","width":1, "color":{"red":1,"green":1,"blue":1}},
                        "right":  {"style": "SOLID","width":1, "color":{"red":1,"green":1,"blue":1}},
                        "innerHorizontal": {"style":"SOLID","width":1,"color":{"red":1,"green":1,"blue":1}},
                        "innerVertical":   {"style":"SOLID","width":1,"color":{"red":1,"green":1,"blue":1}}
                    }
                })
                

                # (O) 표 부분 점선 
                def add_dotted_borders(r1, r2, c1, c2):
                    """바깥+안쪽 모두 DOTTED"""
                    report_fluxus_requests.append({
                        "updateBorders": {
                            "range": {
                                "sheetId": ws_fluxus_report_id,
                                "startRowIndex": r1,
                                "endRowIndex": r2,
                                "startColumnIndex": c1,
                                "endColumnIndex": c2
                            },
                            "top":    {"style": "DOTTED", "width": 1, "color":{"red":0,"green":0,"blue":0}},
                            "bottom": {"style": "DOTTED", "width": 1, "color":{"red":0,"green":0,"blue":0}},
                            "left":   {"style": "DOTTED", "width": 1, "color":{"red":0,"green":0,"blue":0}},
                            "right":  {"style": "DOTTED", "width": 1, "color":{"red":0,"green":0,"blue":0}},
                            "innerHorizontal": {"style": "DOTTED", "width": 1, "color":{"red":0,"green":0,"blue":0}},
                            "innerVertical":   {"style": "DOTTED", "width": 1, "color":{"red":0,"green":0,"blue":0}}
                        }
                    })
                # 1번 섹션 A14:G30 => row=13..30, col=0..7
                add_dotted_borders(13, row_cursor_sum1, 1, 6)
                # 2번 섹션 
                add_dotted_borders(row_cursor_album, row_cursor_sum2, 1, 6)
                # 3번 섹션 
                add_dotted_borders(row_cursor_deduction, row_cursor_sum3-1, 1, 6)
                # 4번 섹션 
                add_dotted_borders(row_cursor_rate, row_cursor_sum4+1, 1, 6)
                

                # (P) 시트 외곽 검정 SOLID 
                report_fluxus_requests.append({
                    "updateBorders": {
                        "range": {
                            "sheetId": ws_fluxus_report_id,
                            "startRowIndex": 0,
                            "endRowIndex": row_cursor_report_end,
                            "startColumnIndex": 0,
                            "endColumnIndex": 7
                        },
                        "top":    {"style": "SOLID","width":1, "color":{"red":0,"green":0,"blue":0}},
                        "bottom": {"style": "SOLID","width":1, "color":{"red":0,"green":0,"blue":0}},
                        "left":   {"style": "SOLID","width":1, "color":{"red":0,"green":0,"blue":0}},
                        "right":  {"style": "SOLID","width":1, "color":{"red":0,"green":0,"blue":0}}
                        # innerHorizontal, innerVertical는 생략 => 기존 값 유지
                    }
                })
                
                all_requests.extend(report_fluxus_requests)

                # batchUpdate 분할 전송
                if len(all_requests) >= 200:
                    sheet_svc.spreadsheets().batchUpdate(
                        spreadsheetId=out_file_id,
                        body={"requests": all_requests}
                    ).execute()
                    all_requests.clear()
                    time.sleep(1)            
            else:
                print(f"소속 코드 오류: {one_sosok}")


    # ---------------------------
    # 마지막으로 남은 요청들을 일괄 처리
    # ---------------------------
    if all_requests:
        sheet_svc.spreadsheets().batchUpdate(
            spreadsheetId=out_file_id,
            body={"requests": all_requests}
        ).execute()
        all_requests.clear()
    time.sleep(1)

    # 루프 끝나면 처리 완료 메시지 (원한다면)
    artist_placeholder.success("모든 아티스트 처리 완료!")

    # ----------------------
    # 다음달 탭 복제 (옵션)
    # ----------------------
    update_next_month_tab(song_cost_sh, ym)
    time.sleep(1)

    # 최종 결과 반환
    return out_file_id


# ========== [5] Streamlit UI =============
def main():
    st.title("아티스트 음원 정산 보고서 자동 생성기")

    # 맨 앞단 - 곡비 파일 제작/수정 섹션
    section_zero_prepare_song_cost()
    st.divider()

    # 섹션 1: 보고서 생성
    section_one_report_input()
    st.divider()

    # 섹션 2: 시트 링크 & 검증
    section_two_sheet_link_and_verification()
    st.divider()

    # 섹션 3: 엑셀 업로드 후 시트분할 ZIP 다운로드
    section_three_upload_and_split_excel()


if __name__ == "__main__":
    main()
